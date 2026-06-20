"""Rebuild verify_progress.json from output file with correct keys."""
import openpyxl, json

# Read original data to get correct row_num
wb_src = openpyxl.load_workbook('sertifikat.xlsx')
ws_src = wb_src.active
all_rows = []
for r in range(2, ws_src.max_row + 1):
    reg = ws_src.cell(r, 3).value
    if reg is None:
        continue
    level = ws_src.cell(r, 5).value
    champ = ws_src.cell(r, 6).value
    org = ws_src.cell(r, 7).value
    name = ws_src.cell(r, 8).value
    url = ws_src.cell(r, 9).value
    ls = str(level).strip() if level else ''
    cs = str(champ).strip() if champ else ''
    os2 = str(org).strip() if org else ''
    ns = str(name).strip() if name else ''
    us = str(url).strip() if url else ''
    if not ls and not cs and not os2 and not ns:
        continue
    if not us or us == 'None':
        continue
    all_rows.append({'row_num': r, 'reg_num': str(reg)})

# Read output file
wb_out = openpyxl.load_workbook('hasil_verifikasi_ai.xlsx')
ws_out = wb_out['Detail Verifikasi']
progress = {}

for idx, src_row in enumerate(all_rows):
    out_r = idx + 2  # output sheet row
    key = f"{src_row['row_num']}_{src_row['reg_num']}"
    status = ws_out.cell(out_r, 12).value or 'NOT_CHECKED'
    notes = ws_out.cell(out_r, 13).value or ''
    method = ws_out.cell(out_r, 3).value or ''
    preview = ws_out.cell(out_r, 14).value or ''
    progress[key] = {
        'status': status,
        'notes': notes,
        'details': {
            'method': method,
            'text_preview': preview,
            'organizer_in_cert': ws_out.cell(out_r, 8).value or '',
            'championship_in_cert': ws_out.cell(out_r, 9).value or '',
            'competition_in_cert': ws_out.cell(out_r, 10).value or '',
            'level_in_cert': ws_out.cell(out_r, 11).value or '',
        },
        'method': method,
    }

with open('verify_progress.json', 'w', encoding='utf-8') as f:
    json.dump(progress, f, ensure_ascii=False, indent=2)

counts = {}
for v in progress.values():
    counts[v['status']] = counts.get(v['status'], 0) + 1
print(f"Rebuilt {len(progress)} entries with correct keys")
for k, v in sorted(counts.items()):
    print(f"  {k}: {v}")
keys = list(progress.keys())[:3]
for k in keys:
    print(f"  sample key: {k}")
