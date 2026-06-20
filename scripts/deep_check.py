"""Deep check: what does the Excel REALLY have vs what the agent reads"""
import openpyxl

wb = openpyxl.load_workbook('sertifikat.xlsx')
ws = wb.active

print(f'Total rows in Excel: {ws.max_row}')
print()

# Categorize every row
categories = {
    'full_data_with_url': [],
    'full_data_no_url': [],
    'reg_only_no_data': [],
    'completely_empty': [],
    'has_some_data_with_url': [],
    'has_some_data_no_url': [],
}

for r in range(2, ws.max_row + 1):
    reg = ws.cell(row=r, column=3).value
    level = ws.cell(row=r, column=5).value
    champ = ws.cell(row=r, column=6).value
    org = ws.cell(row=r, column=7).value
    cert = ws.cell(row=r, column=8).value
    url = ws.cell(row=r, column=9).value
    
    has_reg = reg is not None
    has_url = url is not None and str(url).strip() not in ('', 'None')
    data_fields = [level, champ, org, cert]
    filled_fields = sum(1 for f in data_fields if f is not None and str(f).strip() not in ('', 'None'))
    
    if not has_reg and filled_fields == 0 and not has_url:
        categories['completely_empty'].append(r)
    elif has_reg and filled_fields == 0 and not has_url:
        categories['reg_only_no_data'].append(r)
    elif has_reg and filled_fields > 0 and has_url:
        if filled_fields == 4:
            categories['full_data_with_url'].append(r)
        else:
            categories['has_some_data_with_url'].append(r)
    elif has_reg and filled_fields > 0 and not has_url:
        if filled_fields == 4:
            categories['full_data_no_url'].append(r)
        else:
            categories['has_some_data_no_url'].append(r)
    elif not has_reg and (filled_fields > 0 or has_url):
        print(f'  Row {r}: NO REG but has data! url={str(url)[:50] if has_url else "NO"}')

print('=== CATEGORY SUMMARY ===')
print(f'Complete data + URL (processed): {len(categories["full_data_with_url"])}')
print(f'Some data + URL (processed): {len(categories["has_some_data_with_url"])}')
print(f'Complete data but NO URL: {len(categories["full_data_no_url"])}')
print(f'Some data but NO URL: {len(categories["has_some_data_no_url"])}')
print(f'Reg number only (no cert data, no URL): {len(categories["reg_only_no_data"])}')
print(f'Completely empty: {len(categories["completely_empty"])}')

total_accounted = sum(len(v) for v in categories.values())
print(f'\nTotal accounted: {total_accounted} / {ws.max_row - 1}')

# Show "complete data but no URL" rows
if categories['full_data_no_url']:
    print(f'\n=== COMPLETE DATA BUT NO URL ({len(categories["full_data_no_url"])} rows) ===')
    for r in categories['full_data_no_url'][:20]:
        reg = ws.cell(row=r, column=3).value
        level = ws.cell(row=r, column=5).value
        champ = ws.cell(row=r, column=6).value
        org = ws.cell(row=r, column=7).value
        cert = ws.cell(row=r, column=8).value
        # Check ALL columns for anything URL-like
        all_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() not in ('', 'None'):
                hdr = ws.cell(row=1, column=c).value
                all_vals.append(f'{hdr}={str(v)[:50]}')
        print(f'  Row {r}: {" | ".join(all_vals)}')

# Show "some data + URL" (these have partial data and DO have URL)
if categories['has_some_data_with_url']:
    print(f'\n=== PARTIAL DATA + URL ({len(categories["has_some_data_with_url"])} rows) ===')
    for r in categories['has_some_data_with_url'][:10]:
        reg = ws.cell(row=r, column=3).value
        level = ws.cell(row=r, column=5).value
        champ = ws.cell(row=r, column=6).value
        org = ws.cell(row=r, column=7).value
        cert = ws.cell(row=r, column=8).value
        url = ws.cell(row=r, column=9).value
        print(f'  Row {r}: level={level} champ={champ} org={org} cert={cert} url={str(url)[:50]}')
