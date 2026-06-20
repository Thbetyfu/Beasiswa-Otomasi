"""
AI Certificate Verification Agent v2
=====================================
Two-track approach:
- PDFs: PyMuPDF text extraction + direct text matching (FAST)
- Images: moondream vision model + text matching (SLOWER)

Checks if student-submitted data appears in the actual certificate.
"""

import os
import sys
import json
import time
import base64
import re
import io
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from PIL import Image
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Regex to strip illegal XML characters that openpyxl rejects
ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f'
    r'\ud800-\udfff\ufffe\uffff]'
)

def sanitize(text):
    """Remove illegal characters for Excel/openpyxl."""
    if not isinstance(text, str):
        return text
    return ILLEGAL_CHARACTERS_RE.sub('', text)

# ============================================================
# CONFIGURATION
# ============================================================
OLLAMA_URL = "http://localhost:11434"
VISION_MODEL = "moondream:latest"
CERT_DIR = Path("certs_temp")
PROGRESS_FILE = Path("verify_progress.json")
OUTPUT_FILE = "hasil_verifikasi_ai.xlsx"
TIMEOUT = 120
BATCH_SIZE = 5
MAX_IMAGE_SIZE = 512
WORKERS = 1  # sequential for 6GB VRAM GPU (minicpm-v is 5.5GB)
progress_lock = threading.Lock()

# ============================================================
# PROGRESS / RESUME
# ============================================================

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_progress(progress):
    with open(PROGRESS_FILE, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)

# ============================================================
# DATA LOADING
# ============================================================

def fill_merged_cells(ws):
    """Unmerge cells and fill values from top-left cell to all cells in the range."""
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(start_row=min_row, start_column=min_col, 
                         end_row=max_row, end_column=max_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c, value=top_left_val)

def get_all_rows(xlsx_path):
    """Read sertifikat.xlsx, return all rows with reg_num.
    Rows with cert data + URL: has_cert=True (will be verified)
    Rows with reg_num only (no cert data / no URL): has_cert=False (marked NO_CERTIFICATE)
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    fill_merged_cells(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        reg_num = ws.cell(row=r, column=3).value
        level = ws.cell(row=r, column=5).value
        championship = ws.cell(row=r, column=6).value
        organizer = ws.cell(row=r, column=7).value
        cert_name = ws.cell(row=r, column=8).value
        url = ws.cell(row=r, column=9).value

        if reg_num is None:
            continue

        level_s = str(level).strip() if level else ''
        champ_s = str(championship).strip() if championship else ''
        org_s = str(organizer).strip() if organizer else ''
        name_s = str(cert_name).strip() if cert_name else ''
        url_s = str(url).strip() if url else ''

        has_cert_data = bool(level_s or champ_s or org_s or name_s)
        has_url = bool(url_s and url_s != 'None')

        if has_cert_data and has_url:
            # Row with certificate to verify
            rows.append({
                'row_num': r, 'reg_num': str(reg_num),
                'level': level_s, 'championship': champ_s,
                'organizer': org_s, 'cert_name': name_s, 'url': url_s,
                'has_cert': True,
            })
        else:
            # Row without certificate (hidden row / no submission)
            rows.append({
                'row_num': r, 'reg_num': str(reg_num),
                'level': level_s, 'championship': champ_s,
                'organizer': org_s, 'cert_name': name_s, 'url': url_s if has_url else '',
                'has_cert': False,
            })
    return rows

# ============================================================
# FILE DOWNLOAD
# ============================================================

def download_cert(url, row_num, reg_num):
    ext = '.pdf' if '.pdf' in url.lower() else ('.png' if '.png' in url.lower() else '.jpeg')
    filepath = CERT_DIR / f"row_{row_num}_{reg_num}{ext}"
    if filepath.exists() and filepath.stat().st_size > 0:
        return filepath
    try:
        resp = requests.get(url, timeout=30, stream=True)
        if resp.status_code == 200:
            with open(filepath, 'wb') as f:
                for chunk in resp.iter_content(8192):
                    f.write(chunk)
            return filepath
    except Exception as e:
        pass
    return None

# ============================================================
# TEXT EXTRACTION
# ============================================================

def extract_text_pdf(filepath):
    """Extract text from PDF using PyMuPDF. Returns (text, image_b64_or_None)."""
    try:
        import fitz
        doc = fitz.open(str(filepath))
        all_text = ''
        img_b64 = None
        for page in doc:
            text = page.get_text()
            if text.strip():
                all_text += text + '\n'
            else:
                # Image-based PDF page - render to image
                pix = page.get_pixmap(dpi=200)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                # Resize
                w, h = img.size
                if max(w, h) > MAX_IMAGE_SIZE:
                    ratio = MAX_IMAGE_SIZE / max(w, h)
                    img = img.resize((int(w*ratio), int(h*ratio)), Image.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format='JPEG', quality=85)
                img_b64 = base64.b64encode(buf.getvalue()).decode()
        doc.close()
        return all_text.strip(), img_b64
    except Exception as e:
        return '', None

def extract_text_image(filepath):
    """Prepare image for vision model. Returns image base64."""
    try:
        img = Image.open(filepath).convert('RGB')
        w, h = img.size
        if max(w, h) > MAX_IMAGE_SIZE:
            ratio = MAX_IMAGE_SIZE / max(w, h)
            img = img.resize((int(w*ratio), int(h*ratio)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85)
        return base64.b64encode(buf.getvalue()).decode()
    except:
        return None

def query_vision_model(img_b64):
    """Send image to vision model and ask it to read text."""
    if not img_b64:
        return ''
    try:
        payload = {
            "model": VISION_MODEL,
            "messages": [{
                "role": "user",
                "content": "Baca SEMUA teks pada sertifikat ini dengan teliti. Sebutkan: nama penyelenggara, peringkat/prestasi, nama lomba/kejuaraan, nama peserta, dan tingkat (nasional/internasional/provinsi/kota/kabupaten). Tuliskan persis seperti yang tertera.",
                "images": [img_b64]
            }],
            "stream": False,
            "keep_alive": "15m",
            "options": {"temperature": 0.1, "num_predict": 256, "num_ctx": 2048}
        }
        resp = requests.post(f'{OLLAMA_URL}/api/chat', json=payload, timeout=TIMEOUT)
        if resp.status_code == 200:
            return resp.json().get('message', {}).get('content', '')
    except:
        pass
    return ''

# ============================================================
# VERIFICATION LOGIC
# ============================================================

def normalize(text):
    if not text:
        return ''
    return re.sub(r'\s+', ' ', str(text).strip().lower())

def text_appears(needle, haystack, threshold=0.5):
    """Check if key words from needle appear in haystack.
    Returns True if enough words match."""
    n = normalize(needle)
    h = normalize(haystack)
    if not n or not h:
        return False
    # Direct substring
    if n in h:
        return True
    # Word overlap
    words_n = [w for w in n.split() if len(w) > 2]
    if not words_n:
        return n in h
    matches = sum(1 for w in words_n if w in h)
    return (matches / len(words_n)) >= threshold

def verify_against_text(cert_text, student_data):
    """Check if student data appears in certificate text.
    Returns (status, notes, match_details)."""
    if not cert_text:
        return 'UNREADABLE', 'Tidak ada teks yang bisa diekstrak', {}

    text = normalize(cert_text)
    details = {}
    checks = {}

    # Check organizer
    org = student_data['organizer']
    if org:
        org_match = text_appears(org, text)
        checks['organizer'] = org_match
        details['organizer_in_cert'] = 'YA' if org_match else 'TIDAK'

    # Check championship/achievement
    champ = student_data['championship']
    if champ:
        # Normalize championship keywords
        champ_normalized = champ.lower()
        # Check for key achievement words in the text
        champ_found = False
        for keyword in ['juara', 'harapan', 'finalis', 'final', 'medali', 'emas', 'perak',
                        'perunggu', 'gold', 'silver', 'bronze', 'peserta', 'partisipasi',
                        'winner', 'champion', '1st', '2nd', '3rd', 'first', 'second', 'third']:
            if keyword in champ_normalized and keyword in text:
                champ_found = True
                break
        # Also try direct matching
        if not champ_found:
            champ_found = text_appears(champ, text, threshold=0.4)
        checks['championship'] = champ_found
        details['championship_in_cert'] = 'YA' if champ_found else 'TIDAK'

    # Check competition name
    comp = student_data['cert_name']
    if comp:
        comp_match = text_appears(comp, text, threshold=0.3)
        checks['competition'] = comp_match
        details['competition_in_cert'] = 'YA' if comp_match else 'TIDAK'

    # Check level keywords
    level = student_data['level']
    if level:
        level_keywords = {
            'internasional': ['internasional', 'international', 'asean', 'world', 'global'],
            'nasional': ['nasional', 'national', 'republik indonesia', 'ri ', 'kementrian', 'kementerian'],
            'provinsi': ['provinsi', 'provincial', 'regional', 'wilayah'],
            'kota/kabupaten': ['kota', 'kabupaten', 'kota/kabupaten', 'municipal', 'district'],
        }
        level_norm = normalize(level)
        level_found = False
        for key, keywords in level_keywords.items():
            if key in level_norm:
                level_found = any(kw in text for kw in keywords)
                break
        checks['level'] = level_found
        details['level_in_cert'] = 'YA' if level_found else 'TIDAK'

    # Determine overall status
    if not checks:
        return 'UNREADABLE', 'Tidak ada data untuk diverifikasi', details

    passed = sum(1 for v in checks.values() if v)
    total = len(checks)
    ratio = passed / total

    notes_parts = []
    for field, match in checks.items():
        if not match:
            notes_parts.append(f"{field}: TIDAK DITEMUKAN di sertifikat")

    if ratio >= 0.75:
        status = 'TERVERIFIKASI'
    elif ratio >= 0.5:
        status = 'PARTIAL_MATCH'
    elif ratio > 0:
        status = 'MISMATCH'
    else:
        status = 'MISMATCH'

    notes = f"{passed}/{total} field cocok"
    if notes_parts:
        notes += '; ' + '; '.join(notes_parts)

    return status, notes, details

# ============================================================
# MAIN AGENT
# ============================================================

def run():
    CERT_DIR.mkdir(exist_ok=True)

    print("=" * 60)
    print("  AI CERTIFICATE VERIFICATION AGENT v2")
    print(f"  Vision: {VISION_MODEL} | Text: PyMuPDF + pattern matching")
    print("=" * 60)

    # Check Ollama
    print("\n[1/4] Checking Ollama...")
    try:
        r = requests.get(f'{OLLAMA_URL}/api/tags', timeout=5)
        models = [m['name'] for m in r.json().get('models', [])]
        has_vision = VISION_MODEL in models
        print(f"  Models: {models}")
        print(f"  Vision ({VISION_MODEL}): {'OK' if has_vision else 'MISSING'}")
    except Exception as e:
        print(f"  ERROR: {e}")
        return

    # Load data
    print("\n[2/4] Loading certificate data...")
    all_rows = get_all_rows('sertifikat.xlsx')
    cert_rows = [r for r in all_rows if r.get('has_cert', True)]
    no_cert_rows = [r for r in all_rows if not r.get('has_cert', True)]
    print(f"  Total rows: {len(all_rows)}")
    print(f"  With certificates: {len(cert_rows)} | Without certificate (hidden/empty): {len(no_cert_rows)}")
    pdf_rows = sum(1 for r in cert_rows if '.pdf' in r['url'].lower())
    img_rows = len(cert_rows) - pdf_rows
    print(f"  PDFs: {pdf_rows} | Images: {img_rows}")

    # Load progress
    progress = load_progress()
    remaining = []
    for row in cert_rows:
        key = f"{row['row_num']}_{row['reg_num']}"
        if key not in progress:
            remaining.append(row)
    print(f"  Already done: {len(progress)} | Remaining: {len(remaining)}")

    if not remaining and '--reset' not in sys.argv:
        print("  All done! Use --reset to restart.")
        generate_output(all_rows, progress)
        return

    if '--reset' in sys.argv:
        progress = {}
        remaining = cert_rows

    # Process
    print(f"\n[3/4] Verifying {len(remaining)} certificates...")
    batch_count = 0
    stats = {'TERVERIFIKASI': 0, 'PARTIAL_MATCH': 0, 'MISMATCH': 0, 'UNREADABLE': 0, 'ERROR': 0}
    vision_queue = []  # items needing vision model

    # Phase 1: Download + PDF text extraction (sequential, fast)
    for row in tqdm(remaining, desc="Preparing", unit="cert", ncols=100):
        key = f"{row['row_num']}_{row['reg_num']}"
        is_pdf = '.pdf' in row['url'].lower()

        filepath = download_cert(row['url'], row['row_num'], row['reg_num'])
        if filepath is None:
            with progress_lock:
                progress[key] = {'status': 'ERROR', 'notes': 'Gagal download', 'details': {}, 'method': 'N/A'}
                stats['ERROR'] += 1
            continue

        if is_pdf:
            cert_text, img_b64 = extract_text_pdf(filepath)
            if cert_text:
                # Have text - verify immediately (fast)
                status, notes, details = verify_against_text(cert_text, row)
                details['method'] = 'PDF_text'
                details['text_length'] = len(cert_text)
                details['text_preview'] = cert_text[:200]
                with progress_lock:
                    progress[key] = {'status': status, 'notes': notes, 'details': details, 'method': 'PDF_text'}
                    stats[status] = stats.get(status, 0) + 1
                batch_count += 1
                if batch_count >= BATCH_SIZE:
                    with progress_lock: save_progress(progress)
                    batch_count = 0
                continue
            elif img_b64 and has_vision:
                vision_queue.append((key, row, img_b64, 'PDF_vision'))
                continue
            else:
                with progress_lock:
                    progress[key] = {'status': 'UNREADABLE', 'notes': 'PDF tanpa teks', 'details': {'method': 'PDF_no_text'}, 'method': 'PDF_no_text'}
                    stats['UNREADABLE'] = stats.get('UNREADABLE', 0) + 1
                continue
        else:
            img_b64 = extract_text_image(filepath)
            if img_b64 and has_vision:
                vision_queue.append((key, row, img_b64, 'image'))
            else:
                with progress_lock:
                    progress[key] = {'status': 'UNREADABLE', 'notes': 'Gambar tidak bisa dibaca', 'details': {'method': 'image'}, 'method': 'image'}
                    stats['UNREADABLE'] = stats.get('UNREADABLE', 0) + 1

    with progress_lock: save_progress(progress)
    print(f"\n  PDF text done. {len(vision_queue)} items need vision model...")

    # Phase 2: Vision model queries (concurrent)
    if vision_queue:
        print(f"  Processing {len(vision_queue)} items with {WORKERS} concurrent workers...")

        def process_vision(item):
            key, row, img_b64, method = item
            try:
                ai_text = query_vision_model(img_b64)
                status, notes, details = verify_against_text(ai_text, row)
                details['method'] = method
                details['text_length'] = len(ai_text)
                if ai_text:
                    details['text_preview'] = ai_text[:200]
                return key, status, notes, details, method, None
            except Exception as e:
                return key, 'ERROR', str(e)[:100], {'method': method}, method, e

        done_count = 0
        with ThreadPoolExecutor(max_workers=WORKERS) as executor:
            futures = {executor.submit(process_vision, item): item for item in vision_queue}
            pbar = tqdm(total=len(vision_queue), desc="Vision", unit="cert", ncols=100)
            for future in as_completed(futures):
                key, status, notes, details, method, err = future.result()
                with progress_lock:
                    progress[key] = {'status': status, 'notes': notes, 'details': details, 'method': method}
                    stats[status] = stats.get(status, 0) + 1
                done_count += 1
                pbar.update(1)
                if done_count % BATCH_SIZE == 0:
                    with progress_lock: save_progress(progress)
            pbar.close()

    save_progress(progress)

    # Output
    print(f"\n[4/4] Generating output...\n")
    print("  --- Statistics ---")
    for s, c in sorted(stats.items()):
        print(f"  {s}: {c}")

    generate_output(all_rows, progress)


def generate_output(all_rows, progress):
    wb = openpyxl.Workbook()

    # Styles
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    fills = {
        'TERVERIFIKASI': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'PARTIAL_MATCH': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'MISMATCH': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'UNREADABLE': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        'ERROR': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
        'NO_CERTIFICATE': PatternFill(start_color='DDE6F0', end_color='DDE6F0', fill_type='solid'),
    }

    # Sheet 1: Full Detail
    ws = wb.active
    ws.title = 'Detail Verifikasi'
    headers = ['No', 'Reg Number', 'Method',
               'Level (Data)', 'Championship (Data)', 'Organizer (Data)', 'Cert Name (Data)',
               'Organizer Match', 'Championship Match', 'Competition Match', 'Level Match',
               'Status', 'Notes', 'Text Preview']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    total_stats = {}
    for idx, row in enumerate(all_rows, 1):
        r = idx + 1
        key = f"{row['row_num']}_{row['reg_num']}"
        prog = progress.get(key, {})
        has_cert = row.get('has_cert', True)
        # Rows without cert data get NO_CERTIFICATE status regardless of progress
        status = prog.get('status', 'NOT_CHECKED') if has_cert else 'NO_CERTIFICATE'
        notes = prog.get('notes', '') if has_cert else 'Tidak ada sertifikat yang diunggah'
        details = prog.get('details', {})
        method = prog.get('method', '') if has_cert else 'N/A'

        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=2, value=row['reg_num']).border = border
        ws.cell(row=r, column=3, value=sanitize(method)).border = border
        ws.cell(row=r, column=4, value=sanitize(row['level'])).border = border
        ws.cell(row=r, column=5, value=sanitize(row['championship'])).border = border
        ws.cell(row=r, column=6, value=sanitize(row['organizer'])).border = border
        ws.cell(row=r, column=7, value=sanitize(row['cert_name'])).border = border
        ws.cell(row=r, column=8, value=sanitize(details.get('organizer_in_cert', ''))).border = border
        ws.cell(row=r, column=9, value=sanitize(details.get('championship_in_cert', ''))).border = border
        ws.cell(row=r, column=10, value=sanitize(details.get('competition_in_cert', ''))).border = border
        ws.cell(row=r, column=11, value=sanitize(details.get('level_in_cert', ''))).border = border

        sc = ws.cell(row=r, column=12, value=status)
        sc.border = border; sc.font = Font(bold=True)

        ws.cell(row=r, column=13, value=sanitize(notes)).border = border
        ws.cell(row=r, column=14, value=sanitize(details.get('text_preview', ''))[:200]).border = border

        fill = fills.get(status)
        if fill:
            for c in range(1, 15):
                ws.cell(row=r, column=c).fill = fill

        total_stats[status] = total_stats.get(status, 0) + 1

    widths = [5, 15, 12, 14, 20, 35, 40, 12, 12, 12, 12, 18, 50, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # Sheet 2: Perlu Perhatian (mismatches only)
    ws2 = wb.create_sheet('Perlu Perhatian')
    mm_hdrs = ['No', 'Reg Number', 'Level', 'Championship', 'Organizer',
               'Cert Name', 'Status', 'Notes']
    for c, h in enumerate(mm_hdrs, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border

    mm_r = 2
    mm_i = 1
    for row in all_rows:
        key = f"{row['row_num']}_{row['reg_num']}"
        prog = progress.get(key, {})
        status = prog.get('status', '')
        if status in ['MISMATCH', 'PARTIAL_MATCH']:
            ws2.cell(row=mm_r, column=1, value=mm_i).border = border
            ws2.cell(row=mm_r, column=2, value=sanitize(row['reg_num'])).border = border
            ws2.cell(row=mm_r, column=3, value=sanitize(row['level'])).border = border
            ws2.cell(row=mm_r, column=4, value=sanitize(row['championship'])).border = border
            ws2.cell(row=mm_r, column=5, value=sanitize(row['organizer'])).border = border
            ws2.cell(row=mm_r, column=6, value=sanitize(row['cert_name'])).border = border
            ws2.cell(row=mm_r, column=7, value=status).border = border
            ws2.cell(row=mm_r, column=8, value=sanitize(prog.get('notes', ''))).border = border
            fill = fills.get(status)
            if fill:
                for c in range(1, 9):
                    ws2.cell(row=mm_r, column=c).fill = fill
            mm_r += 1; mm_i += 1

    mm_widths = [5, 15, 14, 20, 35, 40, 16, 50]
    for i, w in enumerate(mm_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3: Statistics
    ws3 = wb.create_sheet('Statistik')
    for c, h in enumerate(['Status', 'Count', 'Percentage'], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border

    total = sum(total_stats.values())
    for i, (s, cnt) in enumerate(sorted(total_stats.items()), 2):
        ws3.cell(row=i, column=1, value=s).border = border
        ws3.cell(row=i, column=2, value=cnt).border = border
        ws3.cell(row=i, column=3, value=f"{cnt/total*100:.1f}%" if total else "0%").border = border
        fill = fills.get(s)
        if fill:
            for c in range(1, 4):
                ws3.cell(row=i, column=c).fill = fill

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 12

    wb.save(OUTPUT_FILE)
    print(f"\n  Output: {OUTPUT_FILE}")
    print(f"  Sheets: Detail Verifikasi, Perlu Perhatian, Statistik")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--reset', action='store_true', help='Reset all progress')
    parser.add_argument('--model', type=str, default=None, help='Override vision model')
    parser.add_argument('--images-only', action='store_true', help='Only process image certs')
    parser.add_argument('--redo-mismatch', action='store_true', help='Re-verify all mismatches')
    args = parser.parse_args()

    if args.model:
        VISION_MODEL = args.model
    if args.reset:
        sys.argv.append('--reset')
    if args.redo_mismatch:
        # Clear progress for mismatched items so they get re-verified
        progress = load_progress()
        keys_to_remove = [k for k, v in progress.items() if v.get('status') in ('MISMATCH', 'UNREADABLE', 'ERROR')]
        for k in keys_to_remove:
            del progress[k]
        save_progress(progress)
        print(f"Cleared {len(keys_to_remove)} mismatch/unreadable/error items for re-verification")

    run()
