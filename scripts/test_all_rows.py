"""Verify that ALL rows from sertifikat.xlsx are in hasil_verifikasi_ai.xlsx"""
import openpyxl

# Read source
src = openpyxl.load_workbook('sertifikat.xlsx').active
print(f"=== SOURCE: sertifikat.xlsx ===")
print(f"Total rows (incl header): {src.max_row}")

# Categorize source rows
has_reg = 0
no_reg = 0
hidden = 0
for r in range(2, src.max_row + 1):
    reg = src.cell(row=r, column=3).value
    is_hidden = src.row_dimensions[r].hidden
    if reg is not None:
        has_reg += 1
    else:
        no_reg += 1
    if is_hidden:
        hidden += 1

print(f"Rows WITH reg_num: {has_reg}")
print(f"Rows WITHOUT reg_num (empty): {no_reg}")
print(f"Hidden rows: {hidden}")

# Read output
out = openpyxl.load_workbook('hasil_verifikasi_ai.xlsx').active
print(f"\n=== OUTPUT: hasil_verifikasi_ai.xlsx ===")
print(f"Total rows (incl header): {out.max_row}")
print(f"Data rows: {out.max_row - 1}")

# Count statuses
from collections import Counter
statuses = Counter()
for r in range(2, out.max_row + 1):
    statuses[out.cell(row=r, column=12).value] += 1

print(f"\nStatus breakdown:")
for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
    print(f"  {s}: {c}")
print(f"  TOTAL: {sum(statuses.values())}")

# Check all reg_nums from source are in output
src_regs = set()
for r in range(2, src.max_row + 1):
    reg = src.cell(row=r, column=3).value
    if reg is not None:
        src_regs.add(str(reg))

out_regs = set()
for r in range(2, out.max_row + 1):
    reg = out.cell(row=r, column=2).value
    if reg is not None:
        out_regs.add(str(reg))

missing = src_regs - out_regs
extra = out_regs - src_regs

print(f"\n=== COVERAGE ===")
print(f"Unique reg_nums in source: {len(src_regs)}")
print(f"Unique reg_nums in output: {len(out_regs)}")
print(f"Missing from output: {len(missing)}")
if missing:
    print(f"  Examples: {list(missing)[:10]}")
print(f"Extra in output: {len(extra)}")

# Final verdict
data_rows = out.max_row - 1
if data_rows == has_reg:
    print(f"\n✓ ALL {has_reg} rows with reg_num are in the output!")
else:
    print(f"\n✗ MISMATCH: source has {has_reg} rows with reg_num, output has {data_rows} data rows")
