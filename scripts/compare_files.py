"""
Compare original Excel vs output to find what user sees as 'bolong'
"""
import openpyxl

# Check original
wb1 = openpyxl.load_workbook('sertifikat.xlsx')
ws1 = wb1.active

# Check output
wb2 = openpyxl.load_workbook('hasil_verifikasi_ai.xlsx')
ws2 = wb2.active

print("=== ORIGINAL sertifikat.xlsx ===")
print(f"Total rows: {ws1.max_row}")
print(f"Hidden rows: 235 (students with no certificate)")
print(f"Visible data rows: {ws1.max_row - 1 - 235} = 755")
print(f"Rows with cert data + URL: 674")
print(f"Rows with nama/reg but NO cert: 235 (hidden)")
print(f"Completely empty rows: 81")
print()

# Count visible non-empty rows user sees in Excel
visible_with_data = 0
visible_reg_nums = set()
for r in range(2, ws1.max_row + 1):
    if ws1.row_dimensions[r].hidden:
        continue
    reg = ws1.cell(row=r, column=3).value
    if reg is not None:
        visible_with_data += 1
        visible_reg_nums.add(str(reg))

print(f"Visible rows with reg_num in Excel: {visible_with_data}")
print(f"Unique visible reg_nums: {len(visible_reg_nums)}")
print()

print("=== OUTPUT hasil_verifikasi_ai.xlsx ===")
print(f"Total rows: {ws2.max_row}")
# Check headers
headers = []
for c in range(1, ws2.max_column + 1):
    h = ws2.cell(row=1, column=c).value
    headers.append(h)
print(f"Headers: {headers}")
print(f"Data rows: {ws2.max_row - 1}")
print()

# Check how many sheets in output
print(f"Output sheets: {wb2.sheetnames}")
for sname in wb2.sheetnames:
    ws = wb2[sname]
    print(f"  Sheet '{sname}': {ws.max_row} rows, {ws.max_column} cols")

# Count unique reg_nums in output
output_reg_nums = set()
for r in range(2, ws2.max_row + 1):
    reg = ws2.cell(row=r, column=1).value  # assuming first col is reg_num
    if reg:
        output_reg_nums.add(str(reg))

print(f"\nUnique reg_nums in output: {len(output_reg_nums)}")

# Find reg_nums in original (visible) but NOT in output
missing = visible_reg_nums - output_reg_nums
print(f"Visible reg_nums NOT in output: {len(missing)}")
if missing:
    for m in sorted(missing)[:10]:
        print(f"  {m}")

# Find reg_nums in output but NOT in original visible
extra = output_reg_nums - visible_reg_nums
print(f"Output reg_nums NOT visible in original: {len(extra)}")

# Check if hidden rows' reg_nums appear in output
hidden_reg_nums = set()
for r in range(2, ws1.max_row + 1):
    if ws1.row_dimensions[r].hidden:
        reg = ws1.cell(row=r, column=3).value
        if reg:
            hidden_reg_nums.add(str(reg))

print(f"\nHidden rows reg_nums: {len(hidden_reg_nums)}")
in_output = hidden_reg_nums & output_reg_nums
print(f"Hidden reg_nums that appear in output: {len(in_output)}")

# Check output sheet 2 and 3
if len(wb2.sheetnames) > 1:
    for sname in wb2.sheetnames[1:]:
        ws = wb2[sname]
        print(f"\n--- Sheet '{sname}' ---")
        # Print first 3 rows
        for r in range(1, min(4, ws.max_row + 1)):
            vals = []
            for c in range(1, min(ws.max_column + 1, 8)):
                v = ws.cell(row=r, column=c).value
                vals.append(str(v)[:30] if v else 'NONE')
            print(f"  Row {r}: {vals}")
