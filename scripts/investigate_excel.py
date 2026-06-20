"""
Investigate why openpyxl may miss data that Microsoft Excel shows.
Check for:
1. Merged cells
2. Hyperlinks vs cell values
3. Hidden rows/columns
4. Different sheet names
5. Cell data types
6. Formulas vs values
7. Rows that have data in col9 but are skipped by current logic
"""
import openpyxl

# Load with data_only=False to see formulas too
wb = openpyxl.load_workbook('sertifikat.xlsx')
ws = wb.active

print(f"Sheet: '{ws.title}'")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print(f"Merged cells: {list(ws.merged_cells.ranges)}")
print()

# Check ALL sheets
print(f"All sheet names: {wb.sheetnames}")
print()

# Check for hidden rows
hidden_rows = []
for r in range(1, ws.max_row + 1):
    if ws.row_dimensions[r].hidden:
        hidden_rows.append(r)
print(f"Hidden rows: {len(hidden_rows)} -> {hidden_rows[:20]}")

# Check for hidden columns
hidden_cols = []
for c in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(c)
    if ws.column_dimensions[col_letter].hidden:
        hidden_cols.append(col_letter)
print(f"Hidden columns: {len(hidden_cols)} -> {hidden_cols}")
print()

# Now check ALL rows that have a URL in column 9
# vs what the current get_all_rows() actually picks up
rows_with_url = []
rows_with_url_but_no_certdata = []
rows_skipped = []

for r in range(2, ws.max_row + 1):
    reg_num = ws.cell(row=r, column=3).value
    level = ws.cell(row=r, column=5).value
    championship = ws.cell(row=r, column=6).value
    organizer = ws.cell(row=r, column=7).value
    cert_name = ws.cell(row=r, column=8).value
    url_cell = ws.cell(row=r, column=9)
    url = url_cell.value
    
    # Also check hyperlink attribute
    hyperlink = url_cell.hyperlink
    
    has_reg = reg_num is not None
    level_s = str(level).strip() if level else ''
    champ_s = str(championship).strip() if championship else ''
    org_s = str(organizer).strip() if organizer else ''
    name_s = str(cert_name).strip() if cert_name else ''
    url_s = str(url).strip() if url else ''
    has_cert_data = bool(level_s or champ_s or org_s or name_s)
    has_url = bool(url_s and url_s != 'None')
    has_hyperlink = hyperlink is not None
    
    if has_url or has_hyperlink:
        info = {
            'row': r, 'reg': str(reg_num)[:20] if reg_num else 'NONE',
            'has_cert': has_cert_data, 'url_val': url_s[:50] if url_s else 'NONE',
            'has_hyperlink': has_hyperlink,
            'hyperlink_target': hyperlink.target[:50] if hyperlink else None,
            'url_type': type(url).__name__,
        }
        rows_with_url.append(info)
        if not has_cert_data:
            rows_with_url_but_no_certdata.append(info)

print(f"=== URL Analysis ===")
print(f"Total rows with URL in col9: {len(rows_with_url)}")
print(f"Rows with URL but NO cert data (col5-8 empty): {len(rows_with_url_but_no_certdata)}")
print()

# Check hyperlink vs value discrepancy
hyperlink_only = [r for r in rows_with_url if r['has_hyperlink'] and not r['url_val']]
value_only = [r for r in rows_with_url if r['url_val'] and not r['has_hyperlink']]
both = [r for r in rows_with_url if r['has_hyperlink'] and r['url_val']]
print(f"Has hyperlink object AND value: {len(both)}")
print(f"Has hyperlink object but NO value: {len(hyperlink_only)}")
print(f"Has value but NO hyperlink object: {len(value_only)}")
print()

if hyperlink_only:
    print("!!! HYPERLINK-ONLY rows (no cell value, but has hyperlink) !!!")
    for item in hyperlink_only[:20]:
        print(f"  Row {item['row']}: reg={item['reg']}, hyperlink={item['hyperlink_target']}")
    if len(hyperlink_only) > 20:
        print(f"  ... and {len(hyperlink_only) - 20} more")
print()

# Now simulate get_all_rows() to see what's skipped
processed = 0
skipped_no_reg = 0
skipped_no_cert = 0
skipped_no_url = 0

for r in range(2, ws.max_row + 1):
    reg_num = ws.cell(row=r, column=3).value
    level = ws.cell(row=r, column=5).value
    championship = ws.cell(row=r, column=6).value
    organizer = ws.cell(row=r, column=7).value
    cert_name = ws.cell(row=r, column=8).value
    url = ws.cell(row=r, column=9).value
    
    if reg_num is None:
        skipped_no_reg += 1
        continue
    level_s = str(level).strip() if level else ''
    champ_s = str(championship).strip() if championship else ''
    org_s = str(organizer).strip() if organizer else ''
    name_s = str(cert_name).strip() if cert_name else ''
    url_s = str(url).strip() if url else ''
    if not level_s and not champ_s and not org_s and not name_s:
        skipped_no_cert += 1
        continue
    if not url_s or url_s == 'None':
        skipped_no_url += 1
        continue
    processed += 1

print(f"=== get_all_rows() simulation ===")
print(f"Processed: {processed}")
print(f"Skipped (no reg_num): {skipped_no_reg}")
print(f"Skipped (no cert data): {skipped_no_cert}")
print(f"Skipped (no URL): {skipped_no_url}")
print(f"Total rows (excl header): {ws.max_row - 1}")
print()

# Check some sample skipped rows in detail
print("=== Sample SKIPPED rows (no cert data) with reg_num ===")
count = 0
for r in range(2, ws.max_row + 1):
    reg_num = ws.cell(row=r, column=3).value
    if reg_num is None:
        continue
    level = ws.cell(row=r, column=5).value
    championship = ws.cell(row=r, column=6).value
    organizer = ws.cell(row=r, column=7).value
    cert_name = ws.cell(row=r, column=8).value
    url = ws.cell(row=r, column=9).value
    level_s = str(level).strip() if level else ''
    champ_s = str(championship).strip() if championship else ''
    org_s = str(organizer).strip() if organizer else ''
    name_s = str(cert_name).strip() if cert_name else ''
    if not level_s and not champ_s and not org_s and not name_s:
        # Print ALL columns for this row
        all_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            all_vals.append(f"c{c}={str(v)[:30] if v else 'NONE'}")
        print(f"  Row {r}: {' | '.join(all_vals)}")
        count += 1
        if count >= 10:
            print(f"  ... showing first 10 of {skipped_no_cert} skipped rows")
            break

# Also check: are there rows where col 4 has data? (maybe data is in different columns)
print()
print("=== Checking if data exists in OTHER columns (1-15) ===")
cols_with_data = {}
for r in range(2, ws.max_row + 1):
    for c in range(1, min(ws.max_column + 1, 16)):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            cols_with_data.setdefault(c, 0)
            cols_with_data[c] += 1

for c in sorted(cols_with_data.keys()):
    col_letter = openpyxl.utils.get_column_letter(c)
    # Get header
    header = ws.cell(row=1, column=c).value
    print(f"  Col {c} ({col_letter}): {cols_with_data[c]} rows have data | Header: {header}")
