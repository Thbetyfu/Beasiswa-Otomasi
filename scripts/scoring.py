import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import re

# ============================================================
# SCORING RULES
# ============================================================

SCORE_TABLE = {
    'Internasional': {
        'juara_1': 15, 'juara_2': 14, 'juara_3': 13,
        'harapan': 12, 'most_inspiring': 11, 'finalis': 10,
    },
    'Nasional': {
        'juara_1': 10, 'juara_2': 9, 'juara_3': 8,
        'harapan': 7, 'most_inspiring': 6, 'finalis': 5,
    },
    'Provinsi': {
        'juara_1': 7, 'juara_2': 6, 'juara_3': 5,
        'harapan': 4, 'most_inspiring': 3, 'finalis': 2,
    },
    'Kota/Kabupaten': {
        'juara_1': 4, 'juara_2': 3, 'juara_3': 2,
        'harapan': 1, 'finalis': 1, 'most_inspiring': 1,
    },
}

# Government organizers at Kota/Kabupaten level (allowed to be scored)
GOVERNMENT_KEYWORDS = [
    'puspresnas', 'pusat prestasi nasional', 'dinas pendidikan',
    'dinas kepemudaan', 'dinas pariwisata', 'kementerian', 'kementrian',
    'kemendikbud', 'kemendikdasmen', 'pemerintah', 'pemda', 'pemkot',
    'pemerintah provinsi', 'pemerintah kota', 'pemerintah kabupaten',
    'badan pengembangan dan pembinaan bahasa', 'kantor bahasa',
    'polri', 'kepolisian', 'tni', 'ipssi', 'ipsi',
    'bpti', 'balai pengembangan talenta',
    'badan kesatuan bangsa', 'suku dinas',
    'disdik', 'dispora', 'disparpora',
    'kejaksaan', 'pengadilan',
]

# Suspicious / paid competition organizers
SUSPICIOUS_ORGANIZERS = [
    'iysa', 'indonesian young scientist association',
    'divya competition', 'divya',
    'goethe institut',
    'pt. bee digital prestasi nusantara', 'pt bee digital',
    'pt. kompetisi online prestasi nusantara', 'pt kompetisi online',
    'pt. sinar sentosa primatama', 'pt sinar sentosa',
    'gypem', 'global youth', 'peace education movement',
    'prestige', 'lembaga prestasi indonesia gemilang',
    'pateron indonesia', 'pateron edukasi',
    'supermachi',
    'aloysius fest',
    'smart student.id', 'smart student',
    'indonesia student news',
    'exsco', 'education expo',
    'fosnas', 'festival olimpiade sains nasional',
    'puskanas', 'pusat kejuaraan sains nasional',
    'gemanessia', 'generasi maju indonesia',
    'lki', 'lembaga kompetisi indonesia', 'olimpiade indonesia',
    'yapresindo',
    'presmas',
    'olympiade sains pelajar',
    'sip publishing',
    'festival olimpiade sains',
    'lembaga kompetisi nasional',
    'pusat kejuaraan',
    'university id education',
    'pt talenta muda bangsa',
    'talenta muda bangsa',
    'posi',
    'bnso',
    'olimnas',
    'olimpiade sains nasional (olimnas)',
    'ilti',
    'ios', 'indonesian olympiad of science',
    'lsp',
]

# E-sport keywords (always 0)
ESPORT_KEYWORDS = [
    'e-sport', 'esport', 'e sport', 'mobile legend', 'mlbb',
    'tekken', 'pubg', 'valorant', 'dota', 'free fire',
    'league of legends', 'genshin', 'fifa', 'pes ',
]

# Non-competition keywords (always 0)
NON_COMPETITION_KEYWORDS = [
    'snbp', 'siswa eligible', 'peserta snbp',
    'dicoding', 'belajar dasar', 'memulai pemrograman',
    'pengenalan data', 'pengenalan ke logika', 'kompetensi kelulusan',
    'kursus bahasa inggris', 'creative', 'lestar komputer',
    'surat tanda selesai belajar',
    'pramuka', 'kwartir', 'kwarcab', 'kwarda', 'jambore',
    'ldks', 'latihan dasar kepemimpinan', 'ldko',
    'osis', 'ketua osis',
    'paskibraka', 'paskibra',
    'magang', 'internship', 'murex resort',
    'tahfidz', 'tahfidz al quran', 'syahadah',
    'jurnalistik', 'reporter',
    'peringkat pertama kelas', 'juara kelas',
    'pelatihan ldko',
    'beauty class', 'wardah',
    'pmr', 'palang merah', 'donor darah',
    'tanda kecakapan',
    'certificate completion', 'completion of',
    'desainer multimedia muda', 'lembaga sertifikasi teknologi',
    'badan nasional sertifikasi profesi',
    'digital marketing',
    'welcoming gen alpha',
    'green youth movement',
    'ieee pre-university', 'workshop on',
    'anbk',
    'cerdas cermat apbn',
    'kompetisi sains ruangguru', 'ruangguru',
    'peserta fls', 'peserta fsl',
    'pemetaan kompetensi matematika',
    'kompetisi sains madrasah',
    'anggota', 'sekertaris', 'organisasi',
    'terpilih sebagai penulis',
    'piagam penghargaan ketua', 'piagam penghargaan sangga',
    'partisipasi', 'partisipas',
]

# Rekognisi Tahfidz - must be > 5 juz
TAHFIDZ_REKOGNISI_MIN_JUZ = 5


def normalize(text):
    """Normalize text for matching: lowercase, strip, remove extra spaces."""
    if not text:
        return ''
    return re.sub(r'\s+', ' ', str(text).strip().lower())


def is_government_organizer(organizer):
    """Check if organizer is a government body."""
    org = normalize(organizer)
    for kw in GOVERNMENT_KEYWORDS:
        if kw in org:
            return True
    return False


def is_suspicious_organizer(organizer):
    """Check if organizer is a known suspicious/paid competition."""
    org = normalize(organizer)
    for kw in SUSPICIOUS_ORGANIZERS:
        if len(kw) <= 4:
            # Short keywords: match as whole word only to avoid false positives
            if re.search(r'\b' + re.escape(kw) + r'\b', org):
                return True
        else:
            if kw in org:
                return True
    return False


def is_esport(name, organizer):
    """Check if the certificate is for e-sports."""
    combined = normalize(name) + ' ' + normalize(organizer)
    for kw in ESPORT_KEYWORDS:
        if kw in combined:
            return True
    return False


def is_non_competition(name, organizer):
    """Check if certificate is non-competitive (course, workshop, org, etc)."""
    combined = normalize(name) + ' ' + normalize(organizer)
    for kw in NON_COMPETITION_KEYWORDS:
        if kw in combined:
            return True
    return False


def is_snbp(name, organizer, championship):
    """Check if it's SNBP-related."""
    combined = normalize(name) + ' ' + normalize(organizer) + ' ' + normalize(championship)
    return 'snbp' in combined or 'siswa eligible' in combined


def get_tahfidz_juz(name):
    """Extract juz number from Tahfidz certificate name. Returns int or None."""
    norm = normalize(name)
    if 'tahfidz' not in norm and 'tahfid' not in norm:
        return None
    # Match patterns like "juz 30 & 29", "10 juz", "5 juzz"
    match = re.search(r'(\d+)\s*(?:&\s*(\d+))?\s*juz', norm)
    if match:
        nums = [int(match.group(1))]
        if match.group(2):
            nums.append(int(match.group(2)))
        return max(nums)
    match = re.search(r'juz\s*(\d+)', norm)
    if match:
        return int(match.group(1))
    return None


def classify_championship(championship, level, organizer, cert_name):
    """
    Classify championship type and return (score_category, flag).
    score_category: 'juara_1', 'juara_2', 'juara_3', 'harapan', 'most_inspiring', 'finalis', 'peserta', 'none'
    flag: string flag or empty
    """
    champ = normalize(championship)
    name = normalize(cert_name)
    org = normalize(organizer)
    flags = []

    # --- ZERO SCORE CHECKS ---

    # E-sport = always 0
    if is_esport(cert_name, organizer):
        return 0, 'E-SPORT = 0'

    # SNBP = always 0
    if is_snbp(cert_name, organizer, championship):
        return 0, 'SNBP = 0'

    # Non-competition certificates = 0
    if is_non_competition(cert_name, organizer):
        # Special check for Tahfidz rekognisi
        juz = get_tahfidz_juz(cert_name)
        if juz is not None and juz > TAHFIDZ_REKOGNISI_MIN_JUZ:
            if level == 'Nasional':
                return 8, 'REKOGNISI Tahfidz (>5 juz)'
            elif level == 'Provinsi':
                return 5, 'REKOGNISI Tahfidz (>5 juz)'
            elif level == 'Internasional':
                return 10, 'REKOGNISI Tahfidz (>5 juz)'
        return 0, 'NON-KOMPETISI = 0'

    # Tahfidz special check even if not caught above
    juz = get_tahfidz_juz(cert_name)
    if juz is not None:
        if juz <= TAHFIDZ_REKOGNISI_MIN_JUZ:
            return 0, f'Tahfidz {juz} juz (<=5) = 0'
        else:
            if level == 'Nasional':
                return 8, 'REKOGNISI Tahfidz (>5 juz)'
            elif level == 'Provinsi':
                return 5, 'REKOGNISI Tahfidz (>5 juz)'
            elif level == 'Internasional':
                return 10, 'REKOGNISI Tahfidz (>5 juz)'
            return 0, 'NON-KOMPETISI = 0'

    # --- CHECK SUSPICIOUS ORGANIZER ---
    if is_suspicious_organizer(organizer):
        flags.append('PERLU VERIFIKASI (penyelenggara mencurigakan)')

    # --- Kota/Kabupaten: only government gets scored ---
    if level == 'Kota/Kabupaten':
        if not is_government_organizer(organizer):
            return 0, 'Kota/Kab non-pemerintah = 0'

    # --- CHAMPIONSHIP CLASSIFICATION ---

    # "Juara Peserta" / just "Peserta" / "Partisipasi"
    if champ in ['juara peserta', 'peserta', 'partisipasi', 'peserta snbp']:
        if level == 'Kota/Kabupaten':
            # Already checked government above, peserta still 0
            return 0, 'Peserta = 0'
        return 0, 'Peserta/Partisipasi = 0'

    # Juara 1
    if re.search(r'juara\s*(1|i\b|satu|utama)', champ) or 'juara 1' in champ or 'juara i ' in champ:
        category = 'juara_1'
        # Special handling: "Juara Utama 2" etc
        if re.search(r'juara\s*(bina\s*2|utama\s*2|madya\s*3|bina\s*3)', champ):
            if 'bina 2' in champ or 'utama 2' in champ:
                category = 'juara_2'
            elif 'madya 3' in champ or 'bina 3' in champ:
                category = 'juara_3'
        if re.search(r'(aktor terbaik|penampilan terbaik)', champ):
            category = 'juara_1'
        score = SCORE_TABLE.get(level, {}).get(category, 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Juara 2
    if re.search(r'juara\s*(2|ii\b|dua|bina\s*2|utama\s*2)', champ):
        score = SCORE_TABLE.get(level, {}).get('juara_2', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Juara 3
    if re.search(r'juara\s*(3|iii\b|tiga|madya\s*3|bina\s*3)', champ):
        score = SCORE_TABLE.get(level, {}).get('juara_3', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Harapan
    if 'harapan' in champ:
        score = SCORE_TABLE.get(level, {}).get('harapan', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Most Inspiring
    if 'most inspiring' in champ or 'penghargaan setara' in champ:
        score = SCORE_TABLE.get(level, {}).get('most_inspiring', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Finalis
    if 'finalis' in champ or 'final' in champ:
        score = SCORE_TABLE.get(level, {}).get('finalis', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Gold Medal / Medali Emas = treat as Juara 1
    if 'gold medal' in champ or 'medali emas' in champ or 'gold medal' in name:
        score = SCORE_TABLE.get(level, {}).get('juara_1', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Silver Medal / Medali Perak = treat as Juara 2
    if 'silver medal' in champ or 'medali perak' in champ or 'medali perak' in name:
        score = SCORE_TABLE.get(level, {}).get('juara_2', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # Bronze Medal / Medali Perunggu = treat as Juara 3
    if 'bronze medal' in champ or 'medali perunggu' in champ:
        score = SCORE_TABLE.get(level, {}).get('juara_3', 0)
        flag_text = '; '.join(flags) if flags else ''
        return score, flag_text

    # If nothing matched
    if flags:
        return 0, 'BELUM DIKLASIFIKASI; ' + '; '.join(flags)
    return 0, 'BELUM DIKLASIFIKASI'


# ============================================================
# MAIN PROCESSING
# ============================================================

def fill_merged_cells(ws):
    """Unmerge cells and fill values from top-left cell to all cells in the range."""
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(start_row=min_row, start_column=min_col, 
                         end_row=max_row, end_column=max_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c, value=top_left_val)


def process():
    wb = openpyxl.load_workbook('sertifikat.xlsx')
    ws = wb.active
    fill_merged_cells(ws)

    # Create output workbook
    out_wb = openpyxl.Workbook()

    # --- Sheet 1: Detail per row ---
    detail_ws = out_wb.active
    detail_ws.title = 'Detail Nilai'

    # Headers
    headers = [
        'No', 'Registration Number', 'Level', 'Championship',
        'Organizer', 'Certificate Name', 'Nilai', 'Keterangan', 'Flag'
    ]
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = detail_ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Color fills
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')

    # Process rows
    student_scores = {}  # reg_number -> { 'scores': [], 'total': 0, 'name': '' }
    row_num = 2
    detail_row = 2

    stats = {
        'total_rows': 0, 'scored': 0, 'zero': 0,
        'flagged': 0, 'unclassified': 0, 'esport': 0,
        'non_competition': 0, 'snbp': 0, 'empty': 0,
    }

    while row_num <= ws.max_row:
        reg_num = ws.cell(row=row_num, column=3).value  # column C = registration_number1
        level = ws.cell(row=row_num, column=5).value      # column E = certificate_level
        championship = ws.cell(row=row_num, column=6).value  # column F
        organizer = ws.cell(row=row_num, column=7).value  # column G
        cert_name = ws.cell(row=row_num, column=8).value  # column H

        if reg_num is None:
            row_num += 1
            continue

        # Check for empty rows (no certificate data at all)
        level_str = str(level).strip() if level else ''
        champ_str = str(championship).strip() if championship else ''
        org_str = str(organizer).strip() if organizer else ''
        name_str = str(cert_name).strip() if cert_name else ''

        if not level_str and not champ_str and not org_str and not name_str:
            stats['total_rows'] += 1
            stats['zero'] += 1
            stats['empty'] += 1
            # Write empty row to detail with note
            detail_ws.cell(row=detail_row, column=1, value=detail_row - 1).border = thin_border
            detail_ws.cell(row=detail_row, column=2, value=str(reg_num)).border = thin_border
            detail_ws.cell(row=detail_row, column=7, value=0).border = thin_border
            detail_ws.cell(row=detail_row, column=8, value='KOSONG (tidak ada sertifikat)').border = thin_border
            gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
            for c in range(1, 10):
                detail_ws.cell(row=detail_row, column=c).fill = gray_fill

            reg_key = str(reg_num)
            if reg_key not in student_scores:
                student_scores[reg_key] = {'scores': [], 'total': 0, 'flagged_count': 0}
            student_scores[reg_key]['scores'].append(0)

            row_num += 1
            detail_row += 1
            continue

        stats['total_rows'] += 1

        # Score the row
        score, flag = classify_championship(champ_str, level_str, org_str, name_str)

        # Update stats
        if score > 0:
            stats['scored'] += 1
        else:
            stats['zero'] += 1

        if 'PERLU VERIFIKASI' in flag:
            stats['flagged'] += 1
        if 'BELUM DIKLASIFIKASI' in flag:
            stats['unclassified'] += 1
        if 'E-SPORT' in flag:
            stats['esport'] += 1
        if 'NON-KOMPETISI' in flag:
            stats['non_competition'] += 1
        if 'SNBP' in flag:
            stats['snbp'] += 1

        # Write to detail sheet
        detail_ws.cell(row=detail_row, column=1, value=detail_row - 1).border = thin_border
        detail_ws.cell(row=detail_row, column=2, value=str(reg_num)).border = thin_border
        detail_ws.cell(row=detail_row, column=3, value=level_str).border = thin_border
        detail_ws.cell(row=detail_row, column=4, value=champ_str).border = thin_border
        detail_ws.cell(row=detail_row, column=5, value=org_str).border = thin_border
        detail_ws.cell(row=detail_row, column=6, value=name_str).border = thin_border

        score_cell = detail_ws.cell(row=detail_row, column=7, value=score)
        score_cell.border = thin_border
        score_cell.alignment = Alignment(horizontal='center')

        detail_ws.cell(row=detail_row, column=8, value=flag).border = thin_border

        # Color code rows
        if score > 0 and 'PERLU VERIFIKASI' not in flag:
            for c in range(1, 10):
                detail_ws.cell(row=detail_row, column=c).fill = green_fill
        elif 'PERLU VERIFIKASI' in flag:
            for c in range(1, 10):
                detail_ws.cell(row=detail_row, column=c).fill = yellow_fill
        elif 'BELUM DIKLASIFIKASI' in flag:
            for c in range(1, 10):
                detail_ws.cell(row=detail_row, column=c).fill = orange_fill
        elif score == 0 and 'E-SPORT' in flag:
            for c in range(1, 10):
                detail_ws.cell(row=detail_row, column=c).fill = red_fill

        # Track student totals
        reg_key = str(reg_num)
        if reg_key not in student_scores:
            student_scores[reg_key] = {'scores': [], 'total': 0, 'flagged_count': 0}
        student_scores[reg_key]['scores'].append(score)
        student_scores[reg_key]['total'] += score
        if 'PERLU VERIFIKASI' in flag:
            student_scores[reg_key]['flagged_count'] += 1

        row_num += 1
        detail_row += 1

    # Set column widths for detail sheet
    col_widths = [5, 18, 16, 22, 40, 50, 8, 40, 30]
    for i, w in enumerate(col_widths, 1):
        detail_ws.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 2: Student Summary ---
    summary_ws = out_wb.create_sheet('Rekap Per Mahasiswa')

    sum_headers = ['No', 'Registration Number', 'Total Nilai', 'Jumlah Sertifikat',
                   'Sertifikat Dinilai', 'Flagged Count', 'Status']
    for col, h in enumerate(sum_headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    sorted_students = sorted(student_scores.items(), key=lambda x: x[1]['total'], reverse=True)

    for idx, (reg_num, data) in enumerate(sorted_students, 1):
        row = idx + 1
        total_cert = len(data['scores'])
        scored_cert = sum(1 for s in data['scores'] if s > 0)
        flagged = data['flagged_count']

        summary_ws.cell(row=row, column=1, value=idx).border = thin_border
        summary_ws.cell(row=row, column=2, value=reg_num).border = thin_border

        total_cell = summary_ws.cell(row=row, column=3, value=data['total'])
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal='center')
        total_cell.font = Font(bold=True, size=11)

        summary_ws.cell(row=row, column=4, value=total_cert).border = thin_border
        summary_ws.cell(row=row, column=5, value=scored_cert).border = thin_border
        summary_ws.cell(row=row, column=6, value=flagged).border = thin_border

        # Status
        if flagged > 0:
            status = 'PERLU VERIFIKASI'
            for c in range(1, 8):
                summary_ws.cell(row=row, column=c).fill = yellow_fill
        elif data['total'] > 0:
            status = 'OK'
            for c in range(1, 8):
                summary_ws.cell(row=row, column=c).fill = green_fill
        else:
            status = 'TIDAK ADA NILAI'
            for c in range(1, 8):
                summary_ws.cell(row=row, column=c).fill = red_fill

        summary_ws.cell(row=row, column=7, value=status).border = thin_border

    # Set column widths for summary
    sum_widths = [5, 18, 14, 18, 18, 15, 20]
    for i, w in enumerate(sum_widths, 1):
        summary_ws.column_dimensions[get_column_letter(i)].width = w

    # --- Sheet 3: Statistics ---
    stats_ws = out_wb.create_sheet('Statistik')
    stats_data = [
        ('Total Baris Diproses', stats['total_rows']),
        ('Sertifikat Dinilai (>0)', stats['scored']),
        ('Sertifikat Nilai 0', stats['zero']),
        ('  - Baris Kosong (no cert)', stats['empty']),
        ('  - E-Sport (=0)', stats['esport']),
        ('  - Non-Kompetisi (=0)', stats['non_competition']),
        ('  - SNBP (=0)', stats['snbp']),
        ('Flagged (Perlu Verifikasi)', stats['flagged']),
        ('Belum Diklasifikasi', stats['unclassified']),
        ('', ''),
        ('Total Mahasiswa', len(student_scores)),
        ('Mahasiswa dengan Nilai > 0', sum(1 for d in student_scores.values() if d['total'] > 0)),
        ('Mahasiswa Flagged', sum(1 for d in student_scores.values() if d['flagged_count'] > 0)),
    ]

    for col, h in enumerate(['Metrik', 'Jumlah'], 1):
        cell = stats_ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for idx, (label, value) in enumerate(stats_data, 2):
        stats_ws.cell(row=idx, column=1, value=label).border = thin_border
        cell = stats_ws.cell(row=idx, column=2, value=value)
        cell.border = thin_border
        cell.font = Font(bold=True)

    stats_ws.column_dimensions['A'].width = 35
    stats_ws.column_dimensions['B'].width = 15

    # Save
    output_file = 'hasil_penilaian_sertifikat.xlsx'
    out_wb.save(output_file)

    print(f"===== HASIL PEMROSESAN =====")
    print(f"Total baris diproses: {stats['total_rows']}")
    print(f"Sertifikat dinilai (>0): {stats['scored']}")
    print(f"Sertifikat nilai 0: {stats['zero']}")
    print(f"  - Baris kosong (no cert): {stats['empty']}")
    print(f"  - E-Sport: {stats['esport']}")
    print(f"  - Non-Kompetisi: {stats['non_competition']}")
    print(f"  - SNBP: {stats['snbp']}")
    print(f"Flagged (Perlu Verifikasi): {stats['flagged']}")
    print(f"Belum Diklasifikasi: {stats['unclassified']}")
    print(f"")
    print(f"Total mahasiswa: {len(student_scores)}")
    print(f"Mahasiswa dengan nilai > 0: {sum(1 for d in student_scores.values() if d['total'] > 0)}")
    print(f"Mahasiswa flagged: {sum(1 for d in student_scores.values() if d['flagged_count'] > 0)}")
    print(f"")
    print(f"Output saved to: {output_file}")
    print(f"Sheets: Detail Nilai, Rekap Per Mahasiswa, Statistik")

    # Print top 10 students
    print(f"\n===== TOP 15 MAHASISWA (by Total Nilai) =====")
    for idx, (reg_num, data) in enumerate(sorted_students[:15], 1):
        flagged_mark = ' [FLAGGED]' if data['flagged_count'] > 0 else ''
        print(f"  {idx}. {reg_num} -> Total: {data['total']}{flagged_mark}")

    # Print unclassified items for review
    print(f"\n===== BELUM DIKLASIFIKASI (perlu manual check) =====")
    detail_row_check = 2
    count = 0
    row_num_check = 2
    while row_num_check <= ws.max_row:
        reg_num = ws.cell(row=row_num_check, column=3).value
        level = ws.cell(row=row_num_check, column=5).value
        championship = ws.cell(row=row_num_check, column=6).value
        organizer = ws.cell(row=row_num_check, column=7).value
        cert_name = ws.cell(row=row_num_check, column=8).value

        if reg_num is None:
            row_num_check += 1
            continue

        level_str = str(level).strip() if level else ''
        champ_str = str(championship).strip() if championship else ''
        org_str = str(organizer).strip() if organizer else ''
        name_str = str(cert_name).strip() if cert_name else ''

        # Skip empty rows
        if not level_str and not champ_str and not org_str and not name_str:
            row_num_check += 1
            continue

        score, flag = classify_championship(champ_str, level_str, org_str, name_str)
        if 'BELUM DIKLASIFIKASI' in flag:
            count += 1
            if count <= 30:
                print(f"  - [{reg_num}] L={level_str} | C={champ_str} | O={org_str} | N={name_str}")
        row_num_check += 1

    if count > 30:
        print(f"  ... and {count - 30} more (check output Excel)")
    if count == 0:
        print("  (None - all items classified!)")


if __name__ == '__main__':
    process()
