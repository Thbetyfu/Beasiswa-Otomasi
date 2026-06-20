"""Check progress and URL analysis"""
import json
import openpyxl

with open('verify_progress.json', 'r', encoding='utf-8') as f:
    progress = json.load(f)

print(f'Total items in progress: {len(progress)}')

statuses = {}
for k, v in progress.items():
    s = v.get('status', 'UNKNOWN')
    statuses[s] = statuses.get(s, 0) + 1

print(f'Status breakdown:')
for s, c in sorted(statuses.items(), key=lambda x: -x[1]):
    print(f'  {s}: {c}')

errors = [(k, v) for k, v in progress.items() if v.get('status') == 'ERROR']
print(f'\nERROR items ({len(errors)}):')
for k, v in errors[:10]:
    print(f'  {k}: {v.get("notes", "")}')

# URL analysis
print(f'\n--- URL analysis ---')
wb = openpyxl.load_workbook('sertifikat.xlsx')
ws = wb.active
url_map = {}
total_with_url = 0
for r in range(2, ws.max_row + 1):
    url = ws.cell(row=r, column=9).value
    reg = ws.cell(row=r, column=3).value
    if url and reg:
        url_s = str(url).strip()
        total_with_url += 1
        if url_s not in url_map:
            url_map[url_s] = []
        url_map[url_s].append(r)

dup_urls = {u: rows for u, rows in url_map.items() if len(rows) > 1}
print(f'Total rows with URL: {total_with_url}')
print(f'Unique URLs: {len(url_map)}')
print(f'URLs shared by multiple rows: {len(dup_urls)}')
for u, rows in list(dup_urls.items())[:5]:
    print(f'  {u[:70]}... -> rows: {rows}')

# Check the 235 "skipped no data" rows more carefully
print(f'\n--- Rows with reg_num but skipped ---')
skipped_rows = []
for r in range(2, ws.max_row + 1):
    reg = ws.cell(row=r, column=3).value
    level = ws.cell(row=r, column=5).value
    champ = ws.cell(row=r, column=6).value
    org = ws.cell(row=r, column=7).value
    cert = ws.cell(row=r, column=8).value
    url = ws.cell(row=r, column=9).value
    
    if reg is None:
        continue
    
    has_data = any([level, champ, org, cert])
    url_s = str(url).strip() if url else ''
    has_url = url_s and url_s != 'None'
    
    if not has_data:
        skipped_rows.append((r, reg, url_s[:60] if has_url else 'NO_URL'))

print(f'Rows with reg_num but no cert data: {len(skipped_rows)}')
for r, reg, url in skipped_rows[:15]:
    print(f'  Row {r}: reg={reg}, url={url}')

# Check output file
print(f'\n--- Output file check ---')
try:
    wb2 = openpyxl.load_workbook('hasil_verifikasi_ai.xlsx')
    ws2 = wb2.active
    print(f'Output sheet: {ws2.title}, rows: {ws2.max_row}')
    
    # Count NOT_CHECKED
    not_checked = 0
    for r in range(2, ws2.max_row + 1):
        status = ws2.cell(row=r, column=12).value
        if status == 'NOT_CHECKED':
            not_checked += 1
    print(f'NOT_CHECKED rows in output: {not_checked}')
except:
    print('Output file not found or error')
