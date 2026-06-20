"""
Certificate Scoring GUI Tool
Dynamic session-based certificate scoring with persistent local storage.
Features: Auto-detect Excel headers, session management, local cert caching, gallery view.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from PIL import Image, ImageTk
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import fitz  # PyMuPDF for PDF rendering
from io import BytesIO
import threading
import os
import json
import shutil
import webbrowser
from datetime import datetime


# ── Helpers ──────────────────────────────────────────────────────────

def get_app_dir():
    """Return the directory where this script lives."""
    return os.path.dirname(os.path.abspath(__file__))


def get_sessions_dir():
    """Return (and create) the sessions root folder."""
    d = os.path.join(get_app_dir(), "sessions")
    os.makedirs(d, exist_ok=True)
    return d


# Header name mapping: possible Excel header names → canonical key
HEADER_ALIASES = {
    'certificate_level': 'level',       'level': 'level',
    'tingkat': 'level',                 'level_certificate': 'level',
    'championship': 'championship',     'keterangan_juara': 'championship',
    'juara': 'championship',            'keterangan juara': 'championship',
    'certificate_organizer': 'organizer', 'organizer': 'organizer',
    'yang_buat_acara': 'organizer',     'yang buat acara': 'organizer',
    'penyelenggara': 'organizer',
    'name': 'cert_name',               'nama_sertifikat': 'cert_name',
    'nama sertifikat': 'cert_name',     'certificate_name': 'cert_name',
    'nama': 'cert_name',
    'url': 'url',                       'link': 'url',
    'link_sertifikat': 'url',           'sertifikat': 'url',
    'link_/_url_sertifikat': 'url',     'link/url_sertifikat': 'url',
    'nilai': 'nilai',                   'score': 'nilai',
    'nilai_sertifikat': 'nilai',
    'ai_status': 'ai_status',           'status_ai': 'ai_status',
    'ai_notes': 'ai_notes',             'keterangan_ai': 'ai_notes',
    'komentar': 'komentar',             'comment': 'komentar',
    'catatan': 'komentar',              'notes': 'komentar',
}

DISPLAY_LABELS = {
    'championship': 'Keterangan Juara',
    'level': 'Level Certificate',
    'organizer': 'Yang Buat Acara',
    'cert_name': 'Nama Sertifikat',
    'url': 'Link',
    'ai_status': 'AI Status',
    'ai_notes': 'AI Notes',
}


class ScoringApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Certificate Scoring Tool")
        self.root.geometry("1200x720")
        self.root.minsize(1000, 600)

        # Data
        self.items = []
        self.current_idx = 0
        self.image_cache = {}     # url -> PIL Image (in-memory for current view)
        self.total_items = 0
        self.scored_count = 0

        # Session
        self.session_dir = None   # path to current session folder
        self.session_name = None

        self._build_ui()
        self._bind_keys()

    # ── UI Construction ──────────────────────────────────────────────

    def _build_ui(self):
        # Top toolbar
        toolbar = ttk.Frame(self.root, padding=8)
        toolbar.pack(fill=tk.X)

        ttk.Button(toolbar, text="Import Excel", command=self._import_excel).pack(side=tk.LEFT, padx=(0, 6))
        self.btn_export = ttk.Button(toolbar, text="Export Excel", command=self._export_excel, state=tk.DISABLED)
        self.btn_export.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_gallery = ttk.Button(toolbar, text="Gallery", command=self._open_gallery, state=tk.DISABLED)
        self.btn_gallery.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_view_data = ttk.Button(toolbar, text="View Data", command=self._open_data_viewer)
        self.btn_view_data.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_sessions = ttk.Button(toolbar, text="Sessions", command=self._open_session_manager)
        self.btn_sessions.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_ai_verify = ttk.Button(toolbar, text="Analisis AI", command=self._run_ai_verify, state=tk.DISABLED)
        self.btn_ai_verify.pack(side=tk.LEFT, padx=(0, 20))

        self.lbl_progress = ttk.Label(toolbar, text="Import or load a session", font=("Segoe UI", 11))
        self.lbl_progress.pack(side=tk.LEFT)

        self.progress_bar = ttk.Progressbar(toolbar, length=200, mode='determinate')
        self.progress_bar.pack(side=tk.LEFT, padx=(12, 0))

        # Main paned area
        paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        # Left: image display
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=3)

        self.img_label = ttk.Label(left_frame, text="Import an Excel file or load a session",
                                   anchor=tk.CENTER, relief=tk.SUNKEN)
        self.img_label.pack(fill=tk.BOTH, expand=True)

        # Right: info + scoring (scrollable)
        right_container = ttk.Frame(paned)
        paned.add(right_container, weight=2)

        right_canvas = tk.Canvas(right_container, highlightthickness=0)
        right_scrollbar = ttk.Scrollbar(right_container, orient=tk.VERTICAL, command=right_canvas.yview)
        right_frame = ttk.Frame(right_canvas, padding=(12, 0, 0, 0))

        right_frame.bind("<Configure>", lambda e: right_canvas.configure(scrollregion=right_canvas.bbox("all")))
        right_canvas.create_window((0, 0), window=right_frame, anchor=tk.NW)
        right_canvas.configure(yscrollcommand=right_scrollbar.set)

        right_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        right_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_right_mousewheel(event):
            right_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_mousewheel(widget):
            """Recursively bind mousewheel to widget and all children."""
            widget.bind("<Enter>", lambda e: right_canvas.bind_all("<MouseWheel>", _on_right_mousewheel))
            widget.bind("<Leave>", lambda e: right_canvas.unbind_all("<MouseWheel>"))
            for child in widget.winfo_children():
                _bind_mousewheel(child)

        # Bind after all children are created (deferred)
        self.root.after(200, lambda: _bind_mousewheel(right_frame))

        # Store reference for later
        self._right_canvas = right_canvas

        # Info fields
        info_frame = ttk.LabelFrame(right_frame, text="Certificate Details", padding=10)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        self.info_vars = {}
        fields = [
            ("Keterangan Juara", "championship"),
            ("Level Certificate", "level"),
            ("Yang Buat Acara", "organizer"),
            ("Nama Sertifikat", "cert_name"),
            ("Link", "url"),
        ]
        for i, (label, key) in enumerate(fields):
            ttk.Label(info_frame, text=f"{label}:", font=("Segoe UI", 9, "bold")).grid(
                row=i, column=0, sticky=tk.W, pady=3)
            var = tk.StringVar()
            self.info_vars[key] = var
            if key == "url":
                lbl = ttk.Label(info_frame, textvariable=var, foreground="blue",
                                cursor="hand2", wraplength=380, justify=tk.LEFT)
                lbl.grid(row=i, column=1, sticky=tk.W, pady=3)
                lbl.bind("<Button-1>", lambda e: self._open_url())
            else:
                ttk.Label(info_frame, textvariable=var, wraplength=380, justify=tk.LEFT).grid(
                    row=i, column=1, sticky=tk.W, pady=3)

        info_frame.columnconfigure(1, weight=1)

        # AI Verification frame
        ai_frame = ttk.LabelFrame(right_frame, text="AI Verification Status", padding=10)
        ai_frame.pack(fill=tk.X, pady=(0, 10))

        self.lbl_ai_status = tk.Label(ai_frame, text="NOT CHECKED", font=("Segoe UI", 10, "bold"),
                                      bg="#9E9E9E", fg="#FFFFFF", padx=8, pady=4)
        self.lbl_ai_status.pack(side=tk.LEFT, padx=(0, 10))

        self.lbl_ai_notes = ttk.Label(ai_frame, text="-", font=("Segoe UI", 9), wraplength=350, justify=tk.LEFT)
        self.lbl_ai_notes.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Scoring section
        score_frame = ttk.LabelFrame(right_frame, text="Scoring", padding=10)
        score_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(score_frame, text="Nilai:", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=0, sticky=tk.W, padx=(0, 8))
        self.nilai_var = tk.StringVar()
        self.nilai_entry = ttk.Entry(score_frame, textvariable=self.nilai_var, width=15,
                                     font=("Segoe UI", 14))
        self.nilai_entry.grid(row=0, column=1, sticky=tk.W)

        ttk.Label(score_frame, text="Komentar:", font=("Segoe UI", 11, "bold")).grid(
            row=1, column=0, sticky=tk.W, padx=(0, 8), pady=(8, 0))
        self.komentar_var = tk.StringVar()
        self.komentar_entry = ttk.Entry(score_frame, textvariable=self.komentar_var, width=35,
                                        font=("Segoe UI", 11))
        self.komentar_entry.grid(row=1, column=1, columnspan=2, sticky=tk.EW, pady=(8, 0))
        score_frame.columnconfigure(1, weight=1)

        btn_row = ttk.Frame(score_frame)
        btn_row.grid(row=2, column=0, columnspan=3, pady=(12, 0))

        self.btn_save = ttk.Button(btn_row, text="Save & Next", command=self._save_and_next, state=tk.DISABLED)
        self.btn_save.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_skip = ttk.Button(btn_row, text="Skip", command=self._skip, state=tk.DISABLED)
        self.btn_skip.pack(side=tk.LEFT, padx=(0, 6))
        self.btn_clear = ttk.Button(btn_row, text="Clear Score", command=self._clear_score, state=tk.DISABLED)
        self.btn_clear.pack(side=tk.LEFT)

        # Navigation section
        nav_frame = ttk.LabelFrame(right_frame, text="Navigation", padding=10)
        nav_frame.pack(fill=tk.X)

        nav_btn_row = ttk.Frame(nav_frame)
        nav_btn_row.pack(fill=tk.X)

        self.btn_first = ttk.Button(nav_btn_row, text="<< First", command=self._go_first, state=tk.DISABLED)
        self.btn_first.pack(side=tk.LEFT, padx=(0, 4))
        self.btn_prev = ttk.Button(nav_btn_row, text="< Prev", command=self._go_prev, state=tk.DISABLED)
        self.btn_prev.pack(side=tk.LEFT, padx=(0, 4))
        self.btn_next = ttk.Button(nav_btn_row, text="Next >", command=self._go_next, state=tk.DISABLED)
        self.btn_next.pack(side=tk.LEFT, padx=(0, 4))
        self.btn_last = ttk.Button(nav_btn_row, text="Last >>", command=self._go_last, state=tk.DISABLED)
        self.btn_last.pack(side=tk.LEFT)

        # Jump to row
        jump_row = ttk.Frame(nav_frame)
        jump_row.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(jump_row, text="Go to:").pack(side=tk.LEFT, padx=(0, 4))
        self.jump_var = tk.StringVar()
        jump_entry = ttk.Entry(jump_row, textvariable=self.jump_var, width=8)
        jump_entry.pack(side=tk.LEFT, padx=(0, 4))
        jump_entry.bind("<Return>", lambda e: self._jump_to())
        ttk.Button(jump_row, text="Go", command=self._jump_to).pack(side=tk.LEFT)

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=(8, 2))
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _bind_keys(self):
        self.root.bind("<Left>", lambda e: self._go_prev())
        self.root.bind("<Right>", lambda e: self._go_next())
        self.root.bind("<Return>", lambda e: self._go_next())
        self.root.bind("<Control-s>", lambda e: self._save_and_next())
        self.root.bind("<Control-e>", lambda e: self._export_excel())

    # ── Dynamic Header Detection ─────────────────────────────────────

    def _detect_headers(self, ws):
        """Scan row 1 and return {canonical_key: column_index} mapping."""
        mapping = {}
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=1, column=col).value
            if raw is None:
                continue
            header = str(raw).strip().lower().replace(' ', '_')
            # Try exact match first, then alias
            canonical = HEADER_ALIASES.get(header)
            if canonical is None:
                # Try without underscores
                header_clean = header.replace('_', '')
                for alias, canon in HEADER_ALIASES.items():
                    if alias.replace('_', '') == header_clean:
                        canonical = canon
                        break
            if canonical and canonical not in mapping:
                mapping[canonical] = col

        return mapping

    # ── Session Management ───────────────────────────────────────────

    def _create_session(self, source_path):
        """Create a new session folder with datetime stamp."""
        now = datetime.now()
        folder_name = now.strftime("%d-%m-%Y_%H-%M")
        session_dir = os.path.join(get_sessions_dir(), folder_name)

        # Handle duplicate folder names (unlikely but safe)
        counter = 1
        while os.path.exists(session_dir):
            folder_name = now.strftime("%d-%m-%Y_%H-%M") + f"_{counter}"
            session_dir = os.path.join(get_sessions_dir(), folder_name)
            counter += 1

        os.makedirs(session_dir, exist_ok=True)
        os.makedirs(os.path.join(session_dir, "certs"), exist_ok=True)

        self.session_dir = session_dir
        self.session_name = folder_name
        return session_dir

    def _save_session(self):
        """Persist current items + scores to session.json."""
        if not self.session_dir or not self.items:
            return
        data = {
            'session_name': self.session_name,
            'created': datetime.now().isoformat(),
            'source_file': getattr(self, '_source_file', ''),
            'total_items': self.total_items,
            'scored_count': sum(1 for it in self.items if it['nilai']),
            'items': self.items,
        }
        path = os.path.join(self.session_dir, "session.json")
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Warning: could not save session: {e}")

    def _load_session(self, session_dir):
        """Load a session from its session.json file."""
        json_path = os.path.join(session_dir, "session.json")
        if not os.path.exists(json_path):
            messagebox.showerror("Error", "session.json not found in this session folder.")
            return False

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read session:\n{e}")
            return False

        self.items = data.get('items', [])
        self.total_items = len(self.items)
        self.current_idx = 0
        self.scored_count = sum(1 for it in self.items if it.get('nilai'))
        self.session_dir = session_dir
        self.session_name = data.get('session_name', os.path.basename(session_dir))
        self._source_file = data.get('source_file', '')
        self.image_cache.clear()

        if not self.items:
            messagebox.showwarning("Empty", "This session has no items.")
            return False

        self._enable_scoring_buttons()
        self.status_var.set(f"Loaded session: {self.session_name} ({self.total_items} items)")
        self._show_current()
        return True

    def _get_all_sessions(self):
        """Return list of session info dicts from the sessions folder."""
        sessions_root = get_sessions_dir()
        result = []
        if not os.path.exists(sessions_root):
            return result

        for name in sorted(os.listdir(sessions_root), reverse=True):
            folder = os.path.join(sessions_root, name)
            if not os.path.isdir(folder):
                continue
            json_path = os.path.join(folder, "session.json")
            info = {
                'name': name,
                'path': folder,
                'total': 0,
                'scored': 0,
                'source': '',
                'created': '',
            }
            if os.path.exists(json_path):
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    info['total'] = data.get('total_items', 0)
                    info['scored'] = data.get('scored_count', 0)
                    info['source'] = os.path.basename(data.get('source_file', ''))
                    info['created'] = data.get('created', '')[:19]
                except Exception:
                    pass
            result.append(info)
        return result

    def _delete_session(self, session_path):
        """Delete an entire session folder."""
        try:
            shutil.rmtree(session_path)
            return True
        except Exception as e:
            messagebox.showerror("Delete Error", f"Could not delete:\n{e}")
            return False

    # ── Import ───────────────────────────────────────────────────────

    REQUIRED_COLUMNS = [
        ('championship', 'Keterangan Juara'),
        ('level',        'Level Certificate'),
        ('organizer',    'Yang Buat Acara'),
        ('cert_name',    'Nama Sertifikat'),
        ('url',          'Link / URL Sertifikat'),
    ]

    def _import_excel(self):
        """Show import dialog: download template or import existing file."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Import Excel")
        dlg.geometry("660x560")
        dlg.resizable(False, True)
        dlg.transient(self.root)
        dlg.grab_set()

        # Centre on parent
        dlg.update_idletasks()
        px = self.root.winfo_x() + (self.root.winfo_width()  - 660) // 2
        py = self.root.winfo_y() + (self.root.winfo_height() - 560) // 2
        dlg.geometry(f"+{max(px,0)}+{max(py,0)}")

        # ── Scrollable container ──────────────────────────────────────
        canvas = tk.Canvas(dlg, highlightthickness=0)
        scrollbar = ttk.Scrollbar(dlg, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        main = ttk.Frame(canvas, padding=22)
        main_win = canvas.create_window((0, 0), window=main, anchor=tk.NW)

        def _on_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        main.bind("<Configure>", _on_configure)

        def _on_canvas_configure(event):
            canvas.itemconfig(main_win, width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        dlg.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))
    
        ttk.Label(main, text="Import Certificate Data",
                  font=("Segoe UI", 16, "bold")).pack(anchor=tk.W, pady=(0, 14))
    
        # ── Option A: Download template ──────────────────────────────
        box_a = ttk.LabelFrame(main, text=" Option A — First time? Download the template ",
                               padding=14)
        box_a.pack(fill=tk.X, pady=(0, 12))
    
        ttk.Label(box_a,
                  text="Download a ready-to-fill Excel template with the correct headers\n"
                       "and an example row. Fill it in, then come back and import it.",
                  font=("Segoe UI", 10), justify=tk.LEFT).pack(anchor=tk.W)
    
        btn_tmpl = ttk.Button(box_a, text="Download Template (.xlsx)",
                              command=lambda: self._download_template(dlg))
        btn_tmpl.pack(anchor=tk.W, pady=(8, 0))
    
        # ── Option B: Import existing file ───────────────────────────
        box_b = ttk.LabelFrame(main, text=" Option B — Already have the file? Import now ",
                               padding=14)
        box_b.pack(fill=tk.X, pady=(0, 12))
    
        ttk.Label(box_b,
                  text="Select an Excel file that follows the required format.\n"
                       "The app will validate it and tell you exactly what is wrong\n"
                       "if the format doesn't match.",
                  font=("Segoe UI", 10), justify=tk.LEFT).pack(anchor=tk.W)
    
        btn_import = ttk.Button(box_b, text="Choose File & Import\u2026",
                                command=lambda: self._do_import_excel(dlg))
        btn_import.pack(anchor=tk.W, pady=(8, 0))
    
        # ── Required columns — shown as individual chips ──────────────
        hint = ttk.LabelFrame(main, text=" Required Columns (Row 1) ", padding=12)
        hint.pack(fill=tk.X)
    
        chip_frame = ttk.Frame(hint)
        chip_frame.pack(fill=tk.X, pady=(0, 8))
    
        for i, (key, lbl) in enumerate(self.REQUIRED_COLUMNS):
            row_i = i // 3
            col_i = i % 3
            chip = tk.Label(chip_frame,
                            text=f"  {lbl}  ",
                            font=("Segoe UI", 9, "bold"),
                            fg="#FFFFFF",
                            bg="#1565C0",
                            relief=tk.FLAT,
                            padx=8, pady=4)
            chip.grid(row=row_i, column=col_i, padx=4, pady=4, sticky=tk.W)
    
        chip_frame.columnconfigure((0, 1, 2), weight=1)
    
        notes_frame = ttk.Frame(hint)
        notes_frame.pack(fill=tk.X)
        notes = [
            "Row 1 = headers  \u2022  Data starts from Row 2",
            "\u2018Link / URL\u2019 column is mandatory (other columns can be empty)",
            "Rows without a URL will be skipped automatically",
        ]
        for n in notes:
            ttk.Label(notes_frame, text=f"\u2022  {n}",
                      font=("Segoe UI", 9), foreground="#555555",
                      justify=tk.LEFT).pack(anchor=tk.W, pady=1)

    # ── Template download ─────────────────────────────────────────────

    def _download_template(self, parent_dlg):
        """Generate and save an Excel template file."""
        path = filedialog.asksaveasfilename(
            title="Save Template As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=get_app_dir(),
            initialfile="template_penilaian_sertifikat.xlsx",
            parent=parent_dlg
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Sertifikat"

            # Write headers with bold styling
            header_style = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF")
            fill_style   = openpyxl.styles.PatternFill(start_color="4472C4",
                                                       end_color="4472C4",
                                                       fill_type="solid")

            for c, (_, lbl) in enumerate(self.REQUIRED_COLUMNS, 1):
                cell = ws.cell(row=1, column=c, value=lbl)
                cell.font = header_style
                cell.fill = fill_style

            # Example row
            example = [
                "Juara 1",
                "Internasional",
                "IEEE",
                "Best Paper Award 2024",
                "https://fileserver.telkomuniversity.ac.id/contoh_sertifikat.pdf",
            ]
            for c, val in enumerate(example, 1):
                ws.cell(row=2, column=c, value=val)

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            wb.save(path)
            messagebox.showinfo("Template Saved",
                                f"Template saved to:\n{path}\n\n"
                                "Fill in the data starting from Row 2,\n"
                                "then use 'Choose File & Import' to load it.",
                                parent=parent_dlg)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save template:\n{e}",
                                 parent=parent_dlg)

    # ── Detailed Excel validation ─────────────────────────────────────

    def _validate_excel_format(self, ws):
        """
        Thoroughly validate an Excel worksheet.
        Returns (is_valid: bool, errors: list[str], warnings: list[str]).
        """
        errors   = []
        warnings = []

        # ── 1. Empty sheet check ────────────────────────────────────
        if ws.max_row is None or ws.max_row < 1:
            errors.append("The sheet is completely empty (no rows at all).")
            return False, errors, warnings

        if ws.max_column is None or ws.max_column < 1:
            errors.append("The sheet has no columns (all cells in Row 1 are blank).")
            return False, errors, warnings

        # ── 2. Header row (Row 1) checks ────────────────────────────
        raw_headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            raw_headers.append((col, val))

        # 2a. Completely blank header row
        if all(v is None or str(v).strip() == '' for _, v in raw_headers):
            errors.append(
                "Row 1 (header row) is completely blank.\n"
                "   → The first row must contain column headers.\n"
                "   → Required headers: " +
                ", ".join(lbl for _, lbl in self.REQUIRED_COLUMNS))
            return False, errors, warnings

        # 2b. Detect which required columns are present
        header_map = self._detect_headers(ws)
        missing_required = []
        for canon_key, display_lbl in self.REQUIRED_COLUMNS:
            if canon_key not in header_map:
                missing_required.append(display_lbl)

        if missing_required:
            found = [lbl for k, lbl in self.REQUIRED_COLUMNS if k in header_map]
            errors.append(
                f"Missing required column(s) in Row 1:\n"
                + "\n".join(f"   ✗ {m}" for m in missing_required)
                + "\n\n   Found columns: "
                + (", ".join(found) if found else "(none recognized)")
                + "\n\n   Tip: Download the template to see the correct format."
            )

        # 2c. Duplicate / unrecognised headers (warning only)
        blank_cols = [c for c, v in raw_headers if v is None or str(v).strip() == '']
        if blank_cols and len(blank_cols) < len(raw_headers):
            col_letters = [openpyxl.utils.get_column_letter(c) for c in blank_cols]
            warnings.append(
                f"Row 1 has blank header(s) at column(s): {', '.join(col_letters)}.\n"
                "   → These columns will be ignored.")

        unrecognised = []
        for col, val in raw_headers:
            if val is None or str(val).strip() == '':
                continue
            h = str(val).strip().lower().replace(' ', '_')
            if h not in HEADER_ALIASES and h.replace('_', '') not in [a.replace('_', '') for a in HEADER_ALIASES]:
                col_letter = openpyxl.utils.get_column_letter(col)
                unrecognised.append(f"Column {col_letter}: '{val}'")
        if unrecognised:
            warnings.append(
                "Unrecognised header(s) (will be ignored):\n   "
                + "\n   ".join(unrecognised))

        # ── 3. Data rows (Row 2+) checks ────────────────────────────
        if ws.max_row < 2:
            errors.append(
                "No data rows found.\n"
                "   → Data must start from Row 2 (Row 1 is the header).")
            return False, errors, warnings

        url_col = header_map.get('url')
        total_data_rows = ws.max_row - 1
        empty_url_rows  = []
        all_blank_rows  = []

        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value
                        for c in range(1, ws.max_column + 1)]
            all_blank = all(v is None or str(v).strip() in ('', 'None')
                            for v in row_vals)
            if all_blank:
                all_blank_rows.append(r)
                continue

            if url_col:
                url_val = ws.cell(row=r, column=url_col).value
                if url_val is None or str(url_val).strip() in ('', 'None'):
                    empty_url_rows.append(r)

        if all_blank_rows and not empty_url_rows:
            # Only blank rows, no URL issues — will result in 0 items
            pass

        if empty_url_rows:
            shown = empty_url_rows[:15]
            suffix = (f"\n   … and {len(empty_url_rows) - 15} more"
                      if len(empty_url_rows) > 15 else '')
            warnings.append(
                f"{len(empty_url_rows)} row(s) have an empty URL and will be skipped:\n"
                f"   Rows: {', '.join(str(r) for r in shown)}{suffix}")

        if all_blank_rows:
            shown = all_blank_rows[:15]
            suffix = (f"\n   … and {len(all_blank_rows) - 15} more"
                      if len(all_blank_rows) > 15 else '')
            warnings.append(
                f"{len(all_blank_rows)} completely blank row(s) found:\n"
                f"   Rows: {', '.join(str(r) for r in shown)}{suffix}\n"
                "   → These rows will be ignored.")

        # ── 4. URL format quick-check (first 5 URLs) ────────────────
        if url_col:
            sample_checked = 0
            bad_urls = []
            for r in range(2, min(ws.max_row + 1, 20)):
                val = ws.cell(row=r, column=url_col).value
                if val and str(val).strip() not in ('', 'None'):
                    url_str = str(val).strip()
                    if not url_str.startswith(('http://', 'https://')):
                        bad_urls.append((r, url_str[:60]))
                    sample_checked += 1
            if bad_urls:
                lines = [f"   Row {r}: '{u}'" for r, u in bad_urls]
                warnings.append(
                    f"{len(bad_urls)} URL(s) don't start with http:// or https://:\n"
                    + "\n".join(lines)
                    + "\n   → These may fail to load.")

        # ── 5. Count valid data rows ─────────────────────────────────
        valid_urls = 0
        if url_col:
            for r in range(2, ws.max_row + 1):
                val = ws.cell(row=r, column=url_col).value
                if val and str(val).strip() not in ('', 'None'):
                    valid_urls += 1

        if valid_urls == 0 and not errors:
            errors.append(
                "No rows with a valid URL found.\n"
                f"   → Total data rows scanned: {total_data_rows}\n"
                f"   → Blank rows: {len(all_blank_rows)}\n"
                f"   → Rows with empty URL: {len(empty_url_rows)}\n"
                "   → At least one row must have a URL in the 'Link / URL' column.")

        is_valid = len(errors) == 0
        return is_valid, errors, warnings

    # ── Actual import (after dialog choice) ────────────────────────────

    def _fill_merged_cells(self, ws):
        """Unmerge cells and fill values from top-left cell to all cells in the range."""
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_val = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(start_row=min_row, start_column=min_col, 
                             end_row=max_row, end_column=max_col)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c, value=top_left_val)

    def _do_import_excel(self, parent_dlg):
        """Open file picker, validate, and import."""
        path = filedialog.askopenfilename(
            title="Select Excel file with certificates",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=get_app_dir(),
            parent=parent_dlg
        )
        if not path:
            return

        self.status_var.set("Validating file…")
        self.root.update_idletasks()

        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            # Fill merged cells first to ensure data consistency
            self._fill_merged_cells(ws)

            # Run full validation
            is_valid, errors, warnings = self._validate_excel_format(ws)

            # Show warnings (non-blocking, informational)
            if warnings:
                warn_text = "\n\n".join(warnings)
                messagebox.showwarning(
                    "Import Warnings",
                    f"The file can be loaded, but note the following:\n\n{warn_text}",
                    parent=parent_dlg)

            if not is_valid:
                err_text = "\n\n".join(errors)
                messagebox.showerror(
                    "Invalid Excel Format",
                    f"The file does not match the required format:\n\n"
                    f"{err_text}\n\n"
                    f"Please fix the issues above or download the template "
                    f"to start with the correct format.",
                    parent=parent_dlg)
                self.status_var.set("Import cancelled — format errors")
                return

            # ── Valid: proceed with import ────────────────────────────
            header_map = self._detect_headers(ws)

            items = []
            for r in range(2, ws.max_row + 1):
                def _get(key):
                    col = header_map.get(key)
                    if col is None:
                        return ''
                    val = ws.cell(row=r, column=col).value
                    return str(val).strip() if val is not None and str(val).strip() != 'None' else ''

                url = _get('url')
                if not url:
                    continue

                items.append({
                    'row_num': r,
                    'level': _get('level'),
                    'championship': _get('championship'),
                    'organizer': _get('organizer'),
                    'cert_name': _get('cert_name'),
                    'url': url,
                    'nilai': _get('nilai'),
                    'is_pdf': '.pdf' in url.lower(),
                    'ai_status': _get('ai_status') or '',
                    'ai_notes': _get('ai_notes') or '',
                    'komentar': _get('komentar'),
                })

            if not items:
                messagebox.showwarning("No Data",
                    "No rows with URLs found in the file.\n"
                    "Make sure data starts from Row 2 and the URL column is filled.",
                    parent=parent_dlg)
                self.status_var.set("Ready")
                return

            self.items = items
            self.total_items = len(items)
            self.current_idx = 0
            self.image_cache.clear()
            self.scored_count = sum(1 for it in items if it['nilai'])

            # Create session
            session_dir = self._create_session(path)
            self._source_file = path
            self._save_session()

            self._enable_scoring_buttons()

            self.status_var.set(
                f"Loaded {self.total_items} certs from {os.path.basename(path)} → Session {self.session_name}")
            self._show_current()

            parent_dlg.destroy()

            messagebox.showinfo("Import Successful",
                f"Successfully imported {self.total_items} certificate(s).\n\n"
                f"Session: {self.session_name}\n"
                f"Already scored: {self.scored_count} / {self.total_items}",
                parent=self.root)

        except openpyxl.utils.exceptions.InvalidFileException as e:
            messagebox.showerror("File Error",
                f"This does not appear to be a valid Excel (.xlsx) file.\n\n"
                f"Details: {e}\n\n"
                f"Please select a proper .xlsx file.",
                parent=parent_dlg)
            self.status_var.set("Import failed — invalid file type")

        except Exception as e:
            messagebox.showerror("Import Error",
                f"An unexpected error occurred:\n\n{e}\n\n"
                "If the problem persists, try downloading the template "
                "and re-entering your data.",
                parent=parent_dlg)
            self.status_var.set("Import failed")

    def _enable_scoring_buttons(self):
        for btn in [self.btn_gallery, self.btn_save, self.btn_skip, self.btn_clear,
                    self.btn_first, self.btn_prev, self.btn_next, self.btn_last,
                    self.btn_ai_verify]:
            btn.config(state=tk.NORMAL)
        # Export is controlled by _update_export_state (only when all scored)
        self._update_export_state()

    # ── Local Cert Cache ─────────────────────────────────────────────

    def _cert_filename(self, idx, url):
        """Generate a local filename for a certificate."""
        ext = '.pdf' if '.pdf' in url.lower() else ('.png' if '.png' in url.lower() else '.jpeg')
        return f"cert_{idx + 1}{ext}"

    def _local_cert_path(self, idx):
        """Return the local path to a cached cert file, or None."""
        if not self.session_dir:
            return None
        fname = self._cert_filename(idx, self.items[idx]['url'])
        p = os.path.join(self.session_dir, "certs", fname)
        return p if os.path.exists(p) else None

    def _save_cert_locally(self, idx, data, url):
        """Save downloaded cert data to session folder."""
        if not self.session_dir:
            return
        fname = self._cert_filename(idx, url)
        p = os.path.join(self.session_dir, "certs", fname)
        try:
            with open(p, 'wb') as f:
                f.write(data)
        except Exception:
            pass

    # ── Display ──────────────────────────────────────────────────────

    def _show_current(self):
        if not self.items:
            return

        item = self.items[self.current_idx]

        # Update info fields
        for key in ['championship', 'level', 'organizer', 'cert_name', 'url']:
            self.info_vars[key].set(item.get(key) or '-')

        # Update AI status
        ai_status = item.get('ai_status', '') or 'NOT_CHECKED'
        ai_notes = item.get('ai_notes', '') or '-'
        self._update_ai_status_ui(ai_status, ai_notes)

        # Update nilai and komentar fields
        self.nilai_var.set(item.get('nilai', ''))
        self.komentar_var.set(item.get('komentar', ''))

        # Update progress
        remaining = self.total_items - self.scored_count
        self.lbl_progress.config(
            text=f"Item {self.current_idx + 1} / {self.total_items}   |   Scored: {self.scored_count}   |   Remaining: {remaining}")
        self.progress_bar['maximum'] = self.total_items
        self.progress_bar['value'] = self.scored_count

        # Update export button state
        self._update_export_state()

        # Load image
        self._load_image(item)

    def _load_image(self, item):
        url = item['url']
        idx = self.current_idx

        if url in self.image_cache:
            self._display_image(self.image_cache[url])
            return

        # Check local cache first
        local_path = self._local_cert_path(idx)
        if local_path:
            self._show_text_placeholder("Loading from cache...")
            threading.Thread(target=self._load_from_local, args=(local_path, url), daemon=True).start()
            return

        # Download
        self._show_text_placeholder("Loading..." if not item['is_pdf'] else "Loading PDF...")
        self.status_var.set(f"Downloading {self.current_idx + 1}/{self.total_items}...")
        threading.Thread(target=self._download_image, args=(url, item['is_pdf'], idx), daemon=True).start()

    def _load_from_local(self, local_path, url):
        """Load image from a locally cached file."""
        try:
            if local_path.lower().endswith('.pdf'):
                doc = fitz.open(local_path)
                page = doc[0]
                pix = page.get_pixmap(matrix=fitz.Matrix(150 / 72, 150 / 72))
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                doc.close()
            else:
                img = Image.open(local_path)
                img.load()

            self.image_cache[url] = img
            self.root.after(0, self._display_image, img)
        except Exception as e:
            # Local cache corrupted, try download
            self.root.after(0, self._show_text_placeholder, f"Cache error, re-downloading...\n{e}")
            item = self.items[self.current_idx]
            threading.Thread(target=self._download_image,
                             args=(url, item['is_pdf'], self.current_idx), daemon=True).start()

    def _download_with_retry(self, url, max_retries=2):
        """Download URL content with retry logic and smart error handling.
        Returns (raw_bytes, content_type) or raises a descriptive exception.
        """
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        last_error = None

        for attempt in range(max_retries + 1):
            try:
                resp = requests.get(url, headers=headers, timeout=30, stream=True, verify=True)

                # Handle specific HTTP errors with clear messages
                if resp.status_code == 404:
                    raise requests.exceptions.HTTPError(
                        f"File not found (404). The URL may be broken or expired:\n{url}")
                elif resp.status_code == 403:
                    raise requests.exceptions.HTTPError(
                        f"Access denied (403). The file may require login or permissions:\n{url}")
                elif resp.status_code == 401:
                    raise requests.exceptions.HTTPError(
                        f"Unauthorized (401). Authentication required:\n{url}")
                elif resp.status_code == 500:
                    if attempt < max_retries:
                        import time
                        time.sleep(1)
                        continue
                    raise requests.exceptions.HTTPError(
                        f"Server error (500). Please try again later:\n{url}")
                elif resp.status_code >= 400:
                    raise requests.exceptions.HTTPError(
                        f"HTTP error {resp.status_code}: {resp.reason}\n{url}")

                # Success
                content_type = resp.headers.get('Content-Type', '')
                return resp.content, content_type

            except requests.exceptions.SSLError:
                # Retry without SSL verification
                try:
                    resp = requests.get(url, headers=headers, timeout=30, stream=True, verify=False)
                    resp.raise_for_status()
                    content_type = resp.headers.get('Content-Type', '')
                    return resp.content, content_type
                except Exception as ssl_err:
                    last_error = ssl_err
                    if attempt < max_retries:
                        continue
                    raise requests.exceptions.SSLError(
                        f"SSL certificate error and fallback failed:\n{ssl_err}")

            except requests.exceptions.ConnectionError as e:
                last_error = e
                if attempt < max_retries:
                    import time
                    time.sleep(1)
                    continue
                raise requests.exceptions.ConnectionError(
                    f"Connection failed after {max_retries + 1} attempts. Check your internet connection.\n{url}")

            except requests.exceptions.Timeout:
                last_error = requests.exceptions.Timeout(
                    f"Download timed out after 30 seconds:\n{url}")
                if attempt < max_retries:
                    continue
                raise last_error

            except requests.exceptions.HTTPError:
                raise  # Re-raise HTTP errors immediately (already have clear messages)

            except Exception as e:
                last_error = e
                if attempt < max_retries:
                    import time
                    time.sleep(0.5)
                    continue
                raise

        if last_error:
            raise last_error

    def _download_image(self, url, is_pdf=False, idx=None):
        try:
            raw_data, content_type = self._download_with_retry(url)

            # Detect if it's actually a PDF (by content-type or URL)
            actual_is_pdf = is_pdf or '.pdf' in url.lower() or 'pdf' in (content_type or '').lower()

            if actual_is_pdf:
                doc = fitz.open(stream=raw_data, filetype="pdf")
                page = doc[0]
                pix = page.get_pixmap(matrix=fitz.Matrix(150 / 72, 150 / 72))
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                doc.close()
            else:
                img = Image.open(BytesIO(raw_data))
                img.load()

            self.image_cache[url] = img

            # Save locally for future use
            if idx is not None:
                self._save_cert_locally(idx, raw_data, url)

            self.root.after(0, self._display_image, img)
        except requests.exceptions.HTTPError as e:
            self.root.after(0, self._show_text_placeholder, f"⚠ {e}")
            self.root.after(0, lambda: self.status_var.set("File not accessible"))
        except requests.exceptions.ConnectionError as e:
            self.root.after(0, self._show_text_placeholder, f"⚠ Connection Error\n\n{e}")
            self.root.after(0, lambda: self.status_var.set("Connection failed"))
        except requests.exceptions.Timeout as e:
            self.root.after(0, self._show_text_placeholder, f"⚠ Timeout\n\n{e}")
            self.root.after(0, lambda: self.status_var.set("Download timed out"))
        except Exception as e:
            self.root.after(0, self._show_text_placeholder, f"Preview not available\n\n{e}")
            self.root.after(0, lambda: self.status_var.set("Download failed"))

    def _update_ai_status_ui(self, status, notes):
        status = status.upper()
        if status == 'TERVERIFIKASI':
            self.lbl_ai_status.config(text="TERVERIFIKASI", bg="#4CAF50", fg="#FFFFFF")
        elif status == 'PARTIAL_MATCH':
            self.lbl_ai_status.config(text="PARTIAL MATCH", bg="#FFEB9C", fg="#795548")
        elif status == 'MISMATCH':
            self.lbl_ai_status.config(text="MISMATCH", bg="#F44336", fg="#FFFFFF")
        elif status == 'UNREADABLE':
            self.lbl_ai_status.config(text="UNREADABLE", bg="#9E9E9E", fg="#FFFFFF")
        elif status == 'ERROR':
            self.lbl_ai_status.config(text="ERROR", bg="#FF9800", fg="#FFFFFF")
        elif status == 'NO_CERTIFICATE':
            self.lbl_ai_status.config(text="NO CERTIFICATE", bg="#E0E0E0", fg="#000000")
        else:
            self.lbl_ai_status.config(text="NOT CHECKED", bg="#9E9E9E", fg="#FFFFFF")
        
        self.lbl_ai_notes.config(text=notes)

    def _extract_fields_from_text(self, text):
        text_lower = text.lower()
        level = ""
        if any(kw in text_lower for kw in ['internasional', 'international', 'world', 'global', 'asean']):
            level = "Internasional"
        elif any(kw in text_lower for kw in ['nasional', 'national', 'republik indonesia', 'ri ']):
            level = "Nasional"
        elif any(kw in text_lower for kw in ['provinsi', 'provincial', 'regional', 'wilayah']):
            level = "Provinsi"
        elif any(kw in text_lower for kw in ['kota', 'kabupaten', 'municipal', 'district']):
            level = "Kota/Kabupaten"
            
        championship = ""
        if any(kw in text_lower for kw in ['juara 1', 'juara i ', 'juara i\b', '1st place', 'first place', 'gold medal', 'medali emas']):
            championship = "Juara 1"
        elif any(kw in text_lower for kw in ['juara 2', 'juara ii', '2nd place', 'second place', 'silver medal', 'medali perak']):
            championship = "Juara 2"
        elif any(kw in text_lower for kw in ['juara 3', 'juara iii', '3rd place', 'third place', 'bronze medal', 'medali perunggu']):
            championship = "Juara 3"
        elif 'harapan' in text_lower:
            championship = "Harapan"
        elif 'finalis' in text_lower or 'finalist' in text_lower:
            championship = "Finalis"
        elif any(kw in text_lower for kw in ['peserta', 'participant', 'partisipasi']):
            championship = "Peserta"
            
        return level, championship

    def _show_ai_progress_dialog(self, total):
        dialog = tk.Toplevel(self.root)
        dialog.title("Analisis AI sedang berjalan")
        dialog.geometry("460x200")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        dialog.update_idletasks()
        px = self.root.winfo_x() + (self.root.winfo_width() - 460) // 2
        py = self.root.winfo_y() + (self.root.winfo_height() - 200) // 2
        dialog.geometry(f"+{max(px,0)}+{max(py,0)}")

        dialog.protocol("WM_DELETE_WINDOW", lambda: self._cancel_ai_verify(dialog))

        ttk.Label(dialog, text="Analisis AI & Otomatisasi Nilai", font=("Segoe UI", 12, "bold")).pack(pady=(15, 10))

        self.ai_lbl_status = ttk.Label(dialog, text="Menyiapkan analisis...", font=("Segoe UI", 10))
        self.ai_lbl_status.pack(pady=2)

        self.ai_progress_bar = ttk.Progressbar(dialog, length=380, mode='determinate')
        self.ai_progress_bar.pack(pady=10)

        self.ai_lbl_eta = ttk.Label(dialog, text="Estimasi Waktu Tersisa: Menghitung...", font=("Segoe UI", 9, "italic"))
        self.ai_lbl_eta.pack(pady=2)

        btn_cancel = ttk.Button(dialog, text="Batal (Cancel)", command=lambda: self._cancel_ai_verify(dialog))
        btn_cancel.pack(pady=(10, 0))

        self.ai_progress_dialog = dialog

    def _cancel_ai_verify(self, dialog):
        if messagebox.askyesno("Batalkan Analisis AI", "Apakah Anda yakin ingin membatalkan proses Analisis AI?"):
            self.ai_cancel_requested = True
            try:
                dialog.destroy()
            except Exception:
                pass

    def _run_ai_verify(self):
        if not self.items:
            return
        if messagebox.askyesno("Analisis AI & Otomatisasi Nilai", 
                               "Jalankan Analisis AI dan pengisian nilai otomatis untuk semua item?\n\n"
                               "Proses ini akan mengunduh sertifikat, membaca teks, mendeteksi kecocokan, "
                               "dan mengisi nilai berdasarkan kriteria beasiswa secara otomatis."):
            self.btn_ai_verify.config(state=tk.DISABLED)
            threading.Thread(target=self._ai_verify_worker, daemon=True).start()

    def _show_ollama_download_prompt(self):
        """Prompt user to download Ollama if missing."""
        if messagebox.askyesno("Ollama Tidak Ditemukan",
                               "Layanan AI lokal (Ollama) tidak ditemukan di komputer Anda.\n\n"
                               "Untuk melakukan analisis lokal offline, Anda perlu menginstal Ollama.\n"
                               "Apakah Anda ingin membuka halaman unduhan Ollama sekarang?"):
            webbrowser.open("https://ollama.com/download")

    def _ensure_ollama_running(self):
        """Check if Ollama is running; if not, try to start it in the background."""
        import requests
        import subprocess
        import shutil
        import os
        import time
        ollama_url = "http://localhost:11434"
        try:
            requests.get(f"{ollama_url}/api/tags", timeout=2)
            return True
        except Exception:
            pass

        # Ollama is not responding, try to start it
        ollama_bin = shutil.which("ollama")
        if not ollama_bin:
            # Try default Windows path
            default_path = os.path.expandvars(r"%LOCALAPPDATA%\Programs\Ollama\ollama.exe")
            if os.path.exists(default_path):
                ollama_bin = default_path

        if not ollama_bin:
            self.root.after(0, self._show_ollama_download_prompt)
            return False

        try:
            startupinfo = None
            if os.name == 'nt':
                # Prevent console window from showing on Windows
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = subprocess.SW_HIDE

            subprocess.Popen([ollama_bin, "serve"], 
                             stdout=subprocess.DEVNULL, 
                             stderr=subprocess.DEVNULL,
                             startupinfo=startupinfo)
            
            # Wait up to 10 seconds for the server to start responding
            for _ in range(10):
                time.sleep(1)
                try:
                    requests.get(f"{ollama_url}/api/tags", timeout=2)
                    return True
                except Exception:
                    pass
        except Exception:
            pass

        return False

    def _ensure_model_available(self):
        """Check if moondream is available; if not, pull it via Ollama API and show progress."""
        import requests
        import json
        ollama_url = "http://localhost:11434"
        vision_model = "moondream:latest"
        try:
            r = requests.get(f"{ollama_url}/api/tags", timeout=3)
            models = [m['name'] for m in r.json().get('models', [])]
            if vision_model in models or any('moondream' in m for m in models):
                return True
        except Exception:
            return False

        # Model is missing, let's pull it
        self.status_var.set("Mengunduh model AI moondream (1.7 GB)... Silakan tunggu.")
        try:
            payload = {"name": "moondream", "stream": True}
            r = requests.post(f"{ollama_url}/api/pull", json=payload, stream=True, timeout=600)
            
            for line in r.iter_lines():
                if line:
                    data = json.loads(line)
                    status = data.get("status", "")
                    completed = data.get("completed", 0)
                    total = data.get("total", 0)
                    
                    if total > 0:
                        pct = int((completed / total) * 100)
                        self.status_var.set(f"Mengunduh model AI moondream: {pct}% selesai...")
                    else:
                        self.status_var.set(f"Mengunduh model AI: {status}")
            
            self.status_var.set("Model AI moondream berhasil diunduh!")
            return True
        except Exception as e:
            print(f"Error pulling model: {e}")
            self.status_var.set(f"Gagal mengunduh model: {e}")
            return False

    def _ai_verify_worker(self):
        import sys
        import time
        scripts_dir = os.path.join(get_app_dir(), 'scripts')
        if scripts_dir not in sys.path:
            sys.path.append(scripts_dir)
            
        try:
            import requests
            from verify_agent import extract_text_pdf, extract_text_image, query_vision_model, verify_against_text
            from scoring import classify_championship
        except ImportError as e:
            self.root.after(0, lambda: messagebox.showerror("Import Error", f"Gagal mengimpor modul scripts:\n{e}"))
            self.root.after(0, self._enable_scoring_buttons)
            return

        # Ensure Ollama is active
        self.status_var.set("Memeriksa status layanan AI (Ollama)...")
        if not self._ensure_ollama_running():
            self.status_var.set("Layanan AI lokal (Ollama) tidak aktif.")
            self.root.after(0, self._enable_scoring_buttons)
            return

        # Ensure model is available
        self.status_var.set("Memeriksa kesediaan model AI...")
        if not self._ensure_model_available():
            self.status_var.set("Model AI tidak tersedia.")
            self.root.after(0, self._enable_scoring_buttons)
            return

        ollama_url = "http://localhost:11434"
        vision_model = "moondream:latest"
        has_vision = False
        try:
            r = requests.get(f"{ollama_url}/api/tags", timeout=3)
            models = [m['name'] for m in r.json().get('models', [])]
            has_vision = vision_model in models or any('moondream' in m for m in models)
        except Exception:
            pass

        total = len(self.items)
        self.ai_cancel_requested = False
        start_time = time.time()
        
        self.root.after(0, lambda: self._show_ai_progress_dialog(total))

        self.scored_count = 0
        for idx, item in enumerate(self.items):
            if self.ai_cancel_requested:
                self.root.after(0, lambda: self.status_var.set("Analisis AI dibatalkan oleh pengguna."))
                self.root.after(0, self._enable_scoring_buttons)
                return

            completed = idx
            elapsed = time.time() - start_time
            if completed > 0:
                avg_time = elapsed / completed
                remaining_time = (total - completed) * avg_time
                minutes = int(remaining_time // 60)
                seconds = int(remaining_time % 60)
                eta_text = f"Estimasi Waktu Tersisa: {minutes:02d} menit {seconds:02d} detik"
            else:
                eta_text = "Estimasi Waktu Tersisa: Menghitung..."
                
            status_text = f"Menganalisis baris {idx+1} dari {total}..."
            
            def _update_ui(c=idx+1, s=status_text, e=eta_text):
                if hasattr(self, 'ai_progress_dialog') and self.ai_progress_dialog.winfo_exists():
                    self.ai_progress_bar.config(value=c)
                    self.ai_lbl_status.config(text=s)
                    self.ai_lbl_eta.config(text=e)
                    
            self.root.after(0, _update_ui)
            
            url = item['url']
            is_pdf = item['is_pdf']
            
            local_path = self._local_cert_path(idx)
            if not local_path:
                try:
                    raw_data, content_type = self._download_with_retry(url)
                    self._save_cert_locally(idx, raw_data, url)
                    local_path = self._local_cert_path(idx)
                except Exception as e:
                    item['ai_status'] = 'ERROR'
                    item['ai_notes'] = f"Gagal download: {e}"
                    item['nilai'] = '0'
                    item['komentar'] = f"AI: Gagal mengunduh sertifikat. {e}"
                    self.root.after(0, self._show_current)
                    self._save_session()
                    continue
            
            cert_text = ""
            img_b64 = None
            method = "N/A"
            
            if is_pdf:
                try:
                    cert_text, img_b64 = extract_text_pdf(local_path)
                    method = "PDF_text" if cert_text else "PDF_vision"
                except Exception as e:
                    item['ai_status'] = 'ERROR'
                    item['ai_notes'] = f"Gagal baca PDF: {e}"
                    item['nilai'] = '0'
                    item['komentar'] = f"AI: Gagal membaca PDF. {e}"
                    self.root.after(0, self._show_current)
                    self._save_session()
                    continue
            else:
                try:
                    img_b64 = extract_text_image(local_path)
                    method = "image"
                except Exception as e:
                    item['ai_status'] = 'ERROR'
                    item['ai_notes'] = f"Gagal baca gambar: {e}"
                    item['nilai'] = '0'
                    item['komentar'] = f"AI: Gagal membaca gambar. {e}"
                    self.root.after(0, self._show_current)
                    self._save_session()
                    continue
                    
            if not cert_text and img_b64:
                if has_vision:
                    try:
                        self.root.after(0, lambda i=idx: self.status_var.set(f"Menjalankan model visi untuk item {i+1}..."))
                        cert_text = query_vision_model(img_b64)
                    except Exception as e:
                        item['ai_status'] = 'ERROR'
                        item['ai_notes'] = f"Vision model error: {e}"
                        item['nilai'] = '0'
                        item['komentar'] = f"AI: Model visi error. {e}"
                        self.root.after(0, self._show_current)
                        self._save_session()
                        continue
                else:
                    item['ai_status'] = 'UNREADABLE'
                    item['ai_notes'] = "Ollama model not available"
                    item['nilai'] = '0'
                    item['komentar'] = "AI: Model Ollama tidak tersedia."
                    self.root.after(0, self._show_current)
                    self._save_session()
                    continue
            
            ai_level, ai_championship = "", ""
            if cert_text:
                ai_level, ai_championship = self._extract_fields_from_text(cert_text)

            if not item.get('level') and ai_level:
                item['level'] = ai_level
            if not item.get('championship') and ai_championship:
                item['championship'] = ai_championship

            if cert_text:
                student_data = {
                    'organizer': item.get('organizer', ''),
                    'championship': item.get('championship', ''),
                    'cert_name': item.get('cert_name', ''),
                    'level': item.get('level', ''),
                }
                try:
                    status, notes, details = verify_against_text(cert_text, student_data)
                    item['ai_status'] = status
                    
                    if status in ('TERVERIFIKASI', 'PARTIAL_MATCH'):
                        score, flag = classify_championship(
                            item.get('championship', ''),
                            item.get('level', ''),
                            item.get('organizer', ''),
                            item.get('cert_name', '')
                        )
                        item['nilai'] = str(score)
                        item['ai_notes'] = f"[{method}] {notes} (Skor: {score})"
                        item['komentar'] = f"AI: Terverifikasi. {notes}."
                    else:
                        item['nilai'] = '0'
                        item['ai_notes'] = f"[{method}] Mismatch: {notes}"
                        item['komentar'] = f"AI: Mismatch. {notes}."
                except Exception as e:
                    item['ai_status'] = 'ERROR'
                    item['ai_notes'] = f"Verify logic error: {e}"
                    item['nilai'] = '0'
                    item['komentar'] = f"AI: Error logika verifikasi. {e}"
            else:
                item['ai_status'] = 'UNREADABLE'
                item['ai_notes'] = "No text found in certificate"
                item['nilai'] = '0'
                item['komentar'] = "AI: Teks tidak ditemukan dalam sertifikat."
                
            if item.get('nilai'):
                self.scored_count += 1
                
            self.root.after(0, self._show_current)
            self._save_session()
            
        def _close_dialog():
            if hasattr(self, 'ai_progress_dialog') and self.ai_progress_dialog.winfo_exists():
                self.ai_progress_dialog.destroy()
        self.root.after(0, _close_dialog)

        self.root.after(0, lambda: self.status_var.set("Analisis AI & Otomatisasi Nilai Selesai!"))
        self.root.after(0, self._enable_scoring_buttons)
        self.root.after(0, lambda: messagebox.showinfo("Selesai", 
                                                       "Analisis AI dan otomatisasi nilai telah selesai untuk semua berkas."))

    def _display_image(self, img):
        try:
            self.img_label.update_idletasks()
            max_w = max(self.img_label.winfo_width(), 400)
            max_h = max(self.img_label.winfo_height(), 300)

            w, h = img.size
            ratio = min(max_w / w, max_h / h, 1.0)
            new_w = int(w * ratio)
            new_h = int(h * ratio)

            resized = img.resize((new_w, new_h), Image.LANCZOS)
            self._tk_image = ImageTk.PhotoImage(resized)
            self.img_label.config(image=self._tk_image, text="")
            self.status_var.set(f"Item {self.current_idx + 1}/{self.total_items}")
        except Exception as e:
            self._show_text_placeholder(f"Display error: {e}")

    def _show_text_placeholder(self, text):
        self.img_label.config(image="", text=text, font=("Segoe UI", 14))

    # ── Scoring ──────────────────────────────────────────────────────

    def _validate_score(self, val):
        """Validate score value. Returns (is_valid, error_message)."""
        if val == '':
            return False, "Please enter a nilai (score) before proceeding."
        try:
            num = float(val)
        except ValueError:
            return False, "Nilai must be a number."
        if num < 0 or num > 20:
            return False, "Nilai must be between 0 and 20."
        return True, ""

    def _apply_score(self):
        """Validate and save current score. Returns True if valid/saved."""
        val = self.nilai_var.get().strip()
        ok, msg = self._validate_score(val)
        if not ok:
            messagebox.showwarning("Invalid Score", msg)
            return False

        komentar = self.komentar_var.get().strip()
        if not komentar:
            messagebox.showwarning("Invalid Comment", "Komentar wajib diisi sebelum lanjut.")
            return False

        old_val = self.items[self.current_idx].get('nilai')
        self.items[self.current_idx]['nilai'] = val
        self.items[self.current_idx]['komentar'] = komentar
        if not old_val:
            self.scored_count += 1
        self._save_session()
        self._update_export_state()
        self.status_var.set(f"Saved nilai={val} for item {self.current_idx + 1}")
        return True

    def _update_export_state(self):
        """Enable Export button only when all items are scored."""
        if self.items and self.scored_count >= self.total_items:
            self.btn_export.config(state=tk.NORMAL)
        else:
            self.btn_export.config(state=tk.DISABLED)

    def _save_and_next(self):
        if not self.items:
            return
        if not self._apply_score():
            return
        if self.current_idx < self.total_items - 1:
            self.current_idx += 1
            self._show_current()
        else:
            self._show_current()
            messagebox.showinfo("Done", "You have reached the last item. Score saved.")

    def _skip(self):
        if not self.items:
            return
        if self.current_idx < self.total_items - 1:
            self.current_idx += 1
            self._show_current()

    def _clear_score(self):
        if not self.items:
            return
        old_val = self.items[self.current_idx].get('nilai')
        self.items[self.current_idx]['nilai'] = ''
        self.items[self.current_idx]['komentar'] = ''
        self.nilai_var.set('')
        self.komentar_var.set('')
        if old_val:
            self.scored_count -= 1
        self._save_session()
        self._show_current()

    # ── Navigation ───────────────────────────────────────────────────

    def _go_prev(self):
        if self.items and self.current_idx > 0:
            self.current_idx -= 1
            self._show_current()

    def _go_next(self):
        if not self.items:
            return
        # Auto-save score before moving to next
        val = self.nilai_var.get().strip()
        if val:
            ok, msg = self._validate_score(val)
            if not ok:
                messagebox.showwarning("Invalid Score", msg)
                return
            if not self._apply_score():
                return
        else:
            messagebox.showwarning("No Score", "Please enter a nilai (0-20) before proceeding.")
            return
        if self.current_idx < self.total_items - 1:
            self.current_idx += 1
            self._show_current()
        else:
            self._show_current()
            if self.scored_count >= self.total_items:
                messagebox.showinfo("All Scored!",
                    f"All {self.total_items} items have been scored.\nYou can now Export to Excel.")

    def _go_first(self):
        if self.items:
            self.current_idx = 0
            self._show_current()

    def _go_last(self):
        if self.items:
            self.current_idx = self.total_items - 1
            self._show_current()

    def _jump_to(self):
        if not self.items:
            return
        try:
            target = int(self.jump_var.get())
            if 1 <= target <= self.total_items:
                self.current_idx = target - 1
                self._show_current()
            else:
                messagebox.showinfo("Out of Range", f"Enter a number between 1 and {self.total_items}")
        except ValueError:
            messagebox.showinfo("Invalid", "Enter a valid number")
        self.jump_var.set('')

    def _open_url(self):
        if self.items:
            url = self.items[self.current_idx]['url']
            if url:
                webbrowser.open(url)

    # ── Data Viewer (Spreadsheet-like editor) ─────────────────────

    def _open_data_viewer(self):
        """Open a spreadsheet-like viewer/editor for imported data.
        Shows all rows, highlights empty URLs (red) and incomplete rows (yellow).
        Supports click-to-select, inline editing, and deletion."""
        if not self.items:
            messagebox.showinfo("No Data",
                "No data loaded yet. Import an Excel file first,\n"
                "then use View Data to inspect and edit the rows.")
            return

        win = tk.Toplevel(self.root)
        win.title(f"Data Viewer — {self.total_items} items")
        win.geometry("1100x650")
        win.minsize(900, 500)
        win.transient(self.root)

        # ── Top: stats + filters ──────────────────────────────────────
        top = ttk.Frame(win, padding=8)
        top.pack(fill=tk.X)

        empty_urls = sum(1 for it in self.items if not it.get('url'))
        no_score = sum(1 for it in self.items if not it.get('nilai'))
        no_champ = sum(1 for it in self.items if not it.get('championship'))
        no_level = sum(1 for it in self.items if not it.get('level'))
        no_org   = sum(1 for it in self.items if not it.get('organizer'))
        no_name  = sum(1 for it in self.items if not it.get('cert_name'))

        stats = (
            f"Total: {self.total_items}   |   "
            f"\u274c Empty URL: {empty_urls}   |   "
            f"\u26a0\ufe0f No Score: {no_score}   |   "
            f"No Juara: {no_champ}   |   No Level: {no_level}   |   "
            f"No Acara: {no_org}   |   No Nama: {no_name}"
        )
        ttk.Label(top, text=stats, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)

        ttk.Label(top, text="    ").pack(side=tk.LEFT)
        leg_red = tk.Label(top, bg="#FFCDD2", width=3, relief=tk.FLAT)
        leg_red.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(top, text="Empty URL", font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(0, 8))
        leg_yel = tk.Label(top, bg="#FFF9C4", width=3, relief=tk.FLAT)
        leg_yel.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(top, text="Has Empty Fields", font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(0, 8))
        leg_grn = tk.Label(top, bg="#C8E6C9", width=3, relief=tk.FLAT)
        leg_grn.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(top, text="Complete", font=("Segoe UI", 8)).pack(side=tk.LEFT)

        # Filter
        filt_frame = ttk.Frame(win, padding=(8, 0, 8, 4))
        filt_frame.pack(fill=tk.X)
        ttk.Label(filt_frame, text="Filter:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 4))
        _dv_filter = [tk.StringVar(value="All")]

        def _dv_refresh(*_a):
            _dv_populate(_dv_filter[0].get())

        for fv in ["All", "Empty URL", "No Score", "Incomplete"]:
            ttk.Radiobutton(filt_frame, text=fv, variable=_dv_filter[0],
                            value=fv, command=_dv_refresh).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(filt_frame, text="Refresh",
                   command=lambda: _dv_populate(_dv_filter[0].get())).pack(side=tk.RIGHT)

        # ── Treeview ──────────────────────────────────────────────────
        tree_frame = ttk.Frame(win)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=8)

        cols = ("#", "Juara", "Level", "Acara", "Nama Sertifikat", "URL", "Nilai")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        tree.heading("#", text="#")
        tree.heading("Juara", text="Keterangan Juara")
        tree.heading("Level", text="Level")
        tree.heading("Acara", text="Yang Buat Acara")
        tree.heading("Nama Sertifikat", text="Nama Sertifikat")
        tree.heading("URL", text="URL / Link")
        tree.heading("Nilai", text="Nilai")

        tree.column("#", width=40, anchor=tk.CENTER, stretch=False)
        tree.column("Juara", width=100, stretch=False)
        tree.column("Level", width=90, stretch=False)
        tree.column("Acara", width=130, stretch=False)
        tree.column("Nama Sertifikat", width=160, stretch=False)
        tree.column("URL", width=360)
        tree.column("Nilai", width=55, anchor=tk.CENTER, stretch=False)

        tree.tag_configure("empty_url", background="#FFCDD2", foreground="#B71C1C")
        tree.tag_configure("has_empty", background="#FFF9C4", foreground="#795548")
        tree.tag_configure("complete",  background="#C8E6C9", foreground="#1B5E20")

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        def _dv_populate(filt="All"):
            for row in tree.get_children():
                tree.delete(row)
            for i, item in enumerate(self.items):
                if filt == "Empty URL" and item.get('url'):
                    continue
                if filt == "No Score" and item.get('nilai'):
                    continue
                if filt == "Incomplete" and all(
                    item.get(k) for k in ('url','championship','level','organizer','cert_name','nilai')
                ):
                    continue
                url_v = item.get('url', '') or ''
                nil_v = item.get('nilai', '') or ''
                if not url_v:
                    tag = "empty_url"
                elif any(not item.get(k) for k in ('championship','level','organizer','cert_name','nilai')):
                    tag = "has_empty"
                else:
                    tag = "complete"
                tree.insert("", "end", iid=str(i), values=(
                    i + 1,
                    item.get('championship', '') or '',
                    item.get('level', '') or '',
                    item.get('organizer', '') or '',
                    item.get('cert_name', '') or '',
                    url_v[:90] + ('...' if len(url_v) > 90 else ''),
                    nil_v,
                ), tags=(tag,))

        _dv_populate()

        # ── Edit panel ────────────────────────────────────────────────
        edit_frame = ttk.LabelFrame(win, text=" Edit Selected Row ", padding=8)
        edit_frame.pack(fill=tk.X, padx=8, pady=(0, 8))

        _dv_edits = {}
        _dv_sel = [None]

        ef_cols = [
            ("Juara:", "championship", 0),
            ("Level:", "level", 1),
            ("Acara:", "organizer", 2),
            ("Nama:", "cert_name", 3),
        ]
        for lbl, key, row_i in ef_cols:
            ttk.Label(edit_frame, text=lbl, font=("Segoe UI", 9, "bold")).grid(
                row=row_i, column=0, sticky=tk.W, padx=(0, 4), pady=2)
            sv = tk.StringVar()
            _dv_edits[key] = sv
            ttk.Entry(edit_frame, textvariable=sv, width=38).grid(
                row=row_i, column=1, sticky=tk.EW, padx=(0, 10), pady=2)

        # URL row
        ttk.Label(edit_frame, text="URL:", font=("Segoe UI", 9, "bold")).grid(
            row=4, column=0, sticky=tk.W, padx=(0, 4), pady=2)
        _url_sv = tk.StringVar()
        _dv_edits['url'] = _url_sv
        ttk.Entry(edit_frame, textvariable=_url_sv, width=80).grid(
            row=4, column=1, columnspan=3, sticky=tk.EW, padx=(0, 10), pady=2)

        # Nilai row
        ttk.Label(edit_frame, text="Nilai:", font=("Segoe UI", 9, "bold")).grid(
            row=5, column=0, sticky=tk.W, padx=(0, 4), pady=2)
        _nil_sv = tk.StringVar()
        _dv_edits['nilai'] = _nil_sv
        ttk.Entry(edit_frame, textvariable=_nil_sv, width=12).grid(
            row=5, column=1, sticky=tk.W, padx=(0, 10), pady=2)

        # Komentar row
        ttk.Label(edit_frame, text="Komentar:", font=("Segoe UI", 9, "bold")).grid(
            row=6, column=0, sticky=tk.W, padx=(0, 4), pady=2)
        _kom_sv = tk.StringVar()
        _dv_edits['komentar'] = _kom_sv
        ttk.Entry(edit_frame, textvariable=_kom_sv, width=80).grid(
            row=6, column=1, columnspan=3, sticky=tk.EW, padx=(0, 10), pady=2)

        edit_frame.columnconfigure(1, weight=1)

        # Buttons
        btn_r = ttk.Frame(edit_frame)
        btn_r.grid(row=7, column=0, columnspan=4, pady=(10, 0))

        def _dv_select(_evt=None):
            sel = tree.selection()
            if not sel:
                return
            idx = int(sel[0])
            _dv_sel[0] = idx
            item = self.items[idx]
            for k, sv in _dv_edits.items():
                sv.set(item.get(k, '') or '')

        def _dv_save():
            idx = _dv_sel[0]
            if idx is None:
                messagebox.showinfo("Info", "Select a row first.")
                return
            nil_val = _dv_edits['nilai'].get().strip()
            if nil_val:
                try:
                    n = float(nil_val)
                    if n < 0 or n > 20:
                        raise ValueError("out of range")
                except ValueError:
                    messagebox.showwarning("Invalid Nilai", "Nilai must be a number between 0 and 20.")
                    return
                # Wajib komentar jika ada nilai
                kom_val = _dv_edits['komentar'].get().strip()
                if not kom_val:
                    messagebox.showwarning("Invalid Comment", "Komentar wajib diisi jika ada nilai.")
                    return
            old_nil = self.items[idx].get('nilai', '')
            for k, sv in _dv_edits.items():
                self.items[idx][k] = sv.get().strip()
            new_nil = self.items[idx].get('nilai', '')
            if old_nil and not new_nil:
                self.scored_count -= 1
            elif not old_nil and new_nil:
                self.scored_count += 1
            # Update tree row in place
            item = self.items[idx]
            url_v = item.get('url', '') or ''
            tag = ("empty_url" if not url_v
                   else "has_empty" if any(not item.get(k) for k in ('championship','level','organizer','cert_name','nilai'))
                   else "complete")
            try:
                tree.item(str(idx), values=(
                    idx + 1,
                    item.get('championship','') or '',
                    item.get('level','') or '',
                    item.get('organizer','') or '',
                    item.get('cert_name','') or '',
                    url_v[:90] + ('...' if len(url_v) > 90 else ''),
                    item.get('nilai','') or '',
                ), tags=(tag,))
            except Exception:
                pass
            for sv in _dv_edits.values():
                sv.set('')
            _dv_sel[0] = None
            self._save_session()
            if self.current_idx == idx:
                self._show_current()
            self.status_var.set(f"Row {idx+1} updated.")

        def _dv_cancel():
            for sv in _dv_edits.values():
                sv.set('')
            _dv_sel[0] = None
            tree.selection_remove(tree.selection())

        def _dv_delete():
            idx = _dv_sel[0]
            if idx is None:
                messagebox.showinfo("Info", "Select a row to delete.")
                return
            if not messagebox.askyesno("Delete Row",
                    f"Delete row #{idx+1}?\n\n"
                    f"URL: {self.items[idx].get('url','(empty)')[:70]}"):
                return
            old = self.items[idx]
            if old.get('nilai'):
                self.scored_count -= 1
            self.items.pop(idx)
            self.total_items = len(self.items)
            if self.current_idx >= self.total_items:
                self.current_idx = max(0, self.total_items - 1)
            for sv in _dv_edits.values():
                sv.set('')
            _dv_sel[0] = None
            _dv_populate(_dv_filter[0].get())
            self._save_session()
            self._show_current()
            win.title(f"Data Viewer — {self.total_items} items")

        ttk.Button(btn_r, text="\U0001f4be Save Changes", command=_dv_save).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_r, text="Cancel", command=_dv_cancel).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_r, text="\U0001f5d1 Delete Row", command=_dv_delete).pack(side=tk.LEFT)

        tree.bind("<<TreeviewSelect>>", _dv_select)

    # ── Session Manager (Gallery) ────────────────────────────────────

    def _open_session_manager(self):
        """Open a window showing all saved sessions with load/delete options."""
        sessions = self._get_all_sessions()

        win = tk.Toplevel(self.root)
        win.title("Session Manager")
        win.geometry("850x600")
        win.minsize(700, 400)

        # Header
        header = ttk.Frame(win, padding=10)
        header.pack(fill=tk.X)
        ttk.Label(header, text="Saved Sessions", font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT)
        ttk.Label(header, text=f"({len(sessions)} sessions)", font=("Segoe UI", 11)).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(header, text="Refresh", command=lambda: self._refresh_session_list(win)).pack(side=tk.RIGHT)

        # Scrollable list
        container = ttk.Frame(win)
        container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=canvas.yview)
        inner = ttk.Frame(canvas)

        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        win.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Store reference for refresh
        win._session_inner = inner
        win._session_canvas = canvas

        self._populate_session_list(inner, sessions, win)

        if not sessions:
            ttk.Label(inner, text="No sessions yet. Import an Excel file to create one.",
                      font=("Segoe UI", 12), padding=40).pack()

    def _refresh_session_list(self, win):
        if not win.winfo_exists():
            return
        for w in win._session_inner.winfo_children():
            w.destroy()
        sessions = self._get_all_sessions()
        self._populate_session_list(win._session_inner, sessions, win)

    def _populate_session_list(self, inner, sessions, win):
        for s in sessions:
            card = ttk.Frame(inner, padding=10, relief=tk.GROOVE)
            card.pack(fill=tk.X, padx=5, pady=4)

            # Session info
            info_frame = ttk.Frame(card)
            info_frame.pack(fill=tk.X)

            ttk.Label(info_frame, text=s['name'], font=("Segoe UI", 13, "bold")).pack(anchor=tk.W)

            detail_text = f"Source: {s['source'] or 'unknown'}   |   Items: {s['total']}   |   Scored: {s['scored']}"
            ttk.Label(info_frame, text=detail_text, font=("Segoe UI", 10)).pack(anchor=tk.W)

            if s['created']:
                ttk.Label(info_frame, text=f"Created: {s['created']}", font=("Segoe UI", 9),
                          foreground="gray").pack(anchor=tk.W)

            # Progress bar
            if s['total'] > 0:
                prog_frame = ttk.Frame(card)
                prog_frame.pack(fill=tk.X, pady=(4, 0))
                prog = ttk.Progressbar(prog_frame, length=400, mode='determinate',
                                       maximum=s['total'], value=s['scored'])
                prog.pack(side=tk.LEFT, fill=tk.X, expand=True)
                pct = int(s['scored'] / s['total'] * 100) if s['total'] > 0 else 0
                ttk.Label(prog_frame, text=f"{pct}%", font=("Segoe UI", 10, "bold"),
                          foreground="green" if pct == 100 else "black").pack(side=tk.LEFT, padx=(8, 0))

            # Buttons
            btn_frame = ttk.Frame(card)
            btn_frame.pack(fill=tk.X, pady=(6, 0))

            is_current = self.session_dir and os.path.normpath(s['path']) == os.path.normpath(self.session_dir)

            if is_current:
                ttk.Label(btn_frame, text="✓ Currently loaded", foreground="green",
                          font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
            else:
                ttk.Button(btn_frame, text="Load Session",
                           command=lambda p=s['path'], w=win: self._load_session_and_close(p, w)).pack(
                    side=tk.LEFT, padx=(0, 6))

            ttk.Button(btn_frame, text="Open Folder",
                       command=lambda p=s['path']: os.startfile(p)).pack(side=tk.LEFT, padx=(0, 6))
            ttk.Button(btn_frame, text="Delete",
                       command=lambda p=s['path'], n=s['name'], w=win: self._confirm_delete_session(p, n, w)).pack(
                side=tk.RIGHT)

    def _load_session_and_close(self, session_path, win):
        if self._load_session(session_path):
            win.destroy()

    def _confirm_delete_session(self, session_path, name, win):
        if messagebox.askyesno("Delete Session",
                               f"Are you sure you want to delete session:\n\n{name}\n\n"
                               "This will remove all data including cached certificates."):
            # If deleting current session, clear the UI
            if self.session_dir and os.path.normpath(session_path) == os.path.normpath(self.session_dir):
                self.items = []
                self.total_items = 0
                self.scored_count = 0
                self.session_dir = None
                self.session_name = None
                self.image_cache.clear()
                self.img_label.config(image="", text="Session deleted")
                self.lbl_progress.config(text="No session loaded")
                for btn in [self.btn_export, self.btn_gallery, self.btn_save, self.btn_skip, self.btn_clear,
                            self.btn_first, self.btn_prev, self.btn_next, self.btn_last,
                            self.btn_ai_verify]:
                    btn.config(state=tk.DISABLED)

            if self._delete_session(session_path):
                self._refresh_session_list(win)

    # ── Gallery (Certificate Grid View) ─────────────────────────────

    def _open_gallery(self):
        if not self.items:
            return

        gal = tk.Toplevel(self.root)
        gal.title(f"Certificate Gallery — {self.total_items} items")
        gal.geometry("1100x700")
        gal.minsize(800, 500)
        self._gallery_win = gal

        # Top bar with filter + legend
        top = ttk.Frame(gal, padding=8)
        top.pack(fill=tk.X)

        ttk.Label(top, text="Filter:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 4))
        self._gal_filter = tk.StringVar(value="All")
        for val in ["All", "Scored", "Not Scored"]:
            ttk.Radiobutton(top, text=val, variable=self._gal_filter,
                            value=val, command=self._refresh_gallery).pack(side=tk.LEFT, padx=(0, 8))

        ttk.Label(top, text="   ").pack(side=tk.LEFT)
        legend_scored = tk.Label(top, bg="#4CAF50", width=3, height=1, relief=tk.FLAT)
        legend_scored.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(top, text="Scored", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 10))
        legend_unscored = tk.Label(top, bg="#BDBDBD", width=3, height=1, relief=tk.FLAT)
        legend_unscored.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(top, text="Not Scored", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 10))
        legend_pdf = tk.Label(top, bg="#2196F3", width=3, height=1, relief=tk.FLAT)
        legend_pdf.pack(side=tk.LEFT, padx=(0, 2))
        ttk.Label(top, text="PDF", font=("Segoe UI", 9)).pack(side=tk.LEFT)

        self._gal_count_label = ttk.Label(top, text="", font=("Segoe UI", 10, "bold"))
        self._gal_count_label.pack(side=tk.RIGHT)

        # Scrollable canvas
        container = ttk.Frame(gal)
        container.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=canvas.yview)
        self._gal_canvas = canvas
        self._gal_inner = ttk.Frame(canvas)

        self._gal_inner.bind("<Configure>",
                             lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._gal_inner, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        gal.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        self._gal_thumbnails = {}
        self._gal_thumb_labels = {}
        self._gal_downloading = set()

        self._refresh_gallery()

    def _refresh_gallery(self):
        if not hasattr(self, '_gallery_win') or not self._gallery_win.winfo_exists():
            return

        for w in self._gal_inner.winfo_children():
            w.destroy()

        filter_val = self._gal_filter.get() if hasattr(self, '_gal_filter') else "All"
        filtered = []
        for idx, item in enumerate(self.items):
            if filter_val == "Scored" and not item.get('nilai'):
                continue
            if filter_val == "Not Scored" and item.get('nilai'):
                continue
            filtered.append((idx, item))

        self._gal_count_label.config(text=f"Showing: {len(filtered)} / {self.total_items}")

        cols = 6
        thumb_w, thumb_h = 150, 110
        pad = 6

        for grid_idx, (item_idx, item) in enumerate(filtered):
            row = grid_idx // cols
            col = grid_idx % cols

            cell = tk.Frame(self._gal_inner, width=thumb_w + pad * 2, height=thumb_h + 50,
                            padx=pad, pady=pad)
            cell.grid(row=row, column=col, padx=2, pady=2)
            cell.grid_propagate(False)

            border_color = "#4CAF50" if item.get('nilai') else "#BDBDBD"
            if item.get('is_pdf') and not item.get('nilai'):
                border_color = "#2196F3"

            thumb_frame = tk.Frame(cell, width=thumb_w, height=thumb_h,
                                   highlightbackground=border_color, highlightthickness=3)
            thumb_frame.pack(fill=tk.X)
            thumb_frame.pack_propagate(False)

            url = item['url']
            if url in self._gal_thumbnails:
                lbl = tk.Label(thumb_frame, image=self._gal_thumbnails[url], cursor="hand2")
            elif url in self.image_cache:
                img = self.image_cache[url]
                thumb = img.copy()
                thumb.thumbnail((thumb_w, thumb_h), Image.LANCZOS)
                photo = ImageTk.PhotoImage(thumb)
                self._gal_thumbnails[url] = photo
                lbl = tk.Label(thumb_frame, image=photo, cursor="hand2")
            else:
                type_text = "PDF" if item.get('is_pdf') else "IMG"
                lbl = tk.Label(thumb_frame, text=f"{type_text}\n#{item_idx + 1}",
                               font=("Segoe UI", 10), bg="#F5F5F5", cursor="hand2")
                if url not in self._gal_downloading:
                    self._gal_downloading.add(url)
                    threading.Thread(target=self._download_gallery_thumb,
                                     args=(url, item.get('is_pdf', False), item_idx), daemon=True).start()

            lbl.pack(fill=tk.BOTH, expand=True)
            lbl.bind("<Button-1>", lambda e, i=item_idx: self._gallery_click(i))
            self._gal_thumb_labels[url] = lbl

            name = item.get('cert_name', '')
            name = name[:22] + "..." if len(name) > 22 else name
            name_lbl = ttk.Label(cell, text=name, font=("Segoe UI", 8), wraplength=thumb_w,
                                 justify=tk.CENTER)
            name_lbl.pack(fill=tk.X, pady=(2, 0))

            score_text = f"Nilai: {item['nilai']}" if item.get('nilai') else "Not scored"
            score_lbl = ttk.Label(cell, text=score_text, font=("Segoe UI", 8, "bold"),
                                  foreground="green" if item.get('nilai') else "gray")
            score_lbl.pack(fill=tk.X)

    def _download_gallery_thumb(self, url, is_pdf, item_idx):
        try:
            # Check local cache first
            local_path = self._local_cert_path(item_idx)
            if local_path:
                if local_path.lower().endswith('.pdf'):
                    doc = fitz.open(local_path)
                    page = doc[0]
                    pix = page.get_pixmap(matrix=fitz.Matrix(100 / 72, 100 / 72))
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    doc.close()
                else:
                    img = Image.open(local_path)
                    img.load()
            else:
                raw_data, content_type = self._download_with_retry(url)
                actual_is_pdf = is_pdf or '.pdf' in url.lower() or 'pdf' in (content_type or '').lower()

                if actual_is_pdf:
                    doc = fitz.open(stream=raw_data, filetype="pdf")
                    page = doc[0]
                    pix = page.get_pixmap(matrix=fitz.Matrix(100 / 72, 100 / 72))
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    doc.close()
                else:
                    img = Image.open(BytesIO(raw_data))
                    img.load()

                self.image_cache[url] = img
                self._save_cert_locally(item_idx, raw_data, url)

            thumb = img.copy()
            thumb.thumbnail((150, 110), Image.LANCZOS)

            def _update():
                if not hasattr(self, '_gallery_win') or not self._gallery_win.winfo_exists():
                    return
                photo = ImageTk.PhotoImage(thumb)
                self._gal_thumbnails[url] = photo
                if url in self._gal_thumb_labels:
                    self._gal_thumb_labels[url].config(image=photo, text="")
            self.root.after(0, _update)
        except Exception:
            # Show "Not Available" placeholder in gallery for failed downloads
            def _show_na():
                if not hasattr(self, '_gallery_win') or not self._gallery_win.winfo_exists():
                    return
                if url in self._gal_thumb_labels:
                    self._gal_thumb_labels[url].config(
                        text="N/A", bg="#FFCDD2", fg="#B71C1C",
                        font=("Segoe UI", 9, "bold"))
            self.root.after(0, _show_na)

    def _gallery_click(self, item_idx):
        self.current_idx = item_idx
        self._show_current()
        if hasattr(self, '_gallery_win') and self._gallery_win.winfo_exists():
            self._gallery_win.destroy()

    # ── Export ────────────────────────────────────────────────────────

    def _export_excel(self):
        if not self.items:
            return
        if self.scored_count < self.total_items:
            messagebox.showwarning("Not Ready",
                f"Please score all items before exporting.\n\n"
                f"Scored: {self.scored_count} / {self.total_items}\n"
                f"Remaining: {self.total_items - self.scored_count}")
            return

        # Default save location: session folder or app dir
        default_dir = self.session_dir or get_app_dir()
        default_name = f"hasil_penilaian_{self.session_name or 'export'}.xlsx"

        path = filedialog.asksaveasfilename(
            title="Export Results",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=default_dir,
            initialfile=default_name
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Penilaian Sertifikat"

            headers = ["Keterangan Juara", "Level Certificate", "Yang Buat Acara",
                       "Nama Sertifikat", "Link", "Nilai", "AI Status", "AI Notes", "Komentar"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = openpyxl.styles.Font(bold=True)

            for r, item in enumerate(self.items, 2):
                ws.cell(row=r, column=1, value=item.get('championship', ''))
                ws.cell(row=r, column=2, value=item.get('level', ''))
                ws.cell(row=r, column=3, value=item.get('organizer', ''))
                ws.cell(row=r, column=4, value=item.get('cert_name', ''))
                ws.cell(row=r, column=5, value=item.get('url', ''))
                nilai_val = item.get('nilai', '')
                if nilai_val:
                    try:
                        ws.cell(row=r, column=6, value=float(nilai_val))
                    except ValueError:
                        ws.cell(row=r, column=6, value=nilai_val)
                ws.cell(row=r, column=7, value=item.get('ai_status', ''))
                ws.cell(row=r, column=8, value=item.get('ai_notes', ''))
                ws.cell(row=r, column=9, value=item.get('komentar', ''))

            # Auto-width columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                    except Exception:
                        cell_len = 0
                    if cell_len > max_len:
                        max_len = cell_len
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

            wb.save(path)
            scored = sum(1 for it in self.items if it['nilai'])
            self.status_var.set(f"Exported {scored}/{self.total_items} scored items")
            messagebox.showinfo("Export Complete",
                                f"Saved to:\n{path}\n\nScored: {scored}/{self.total_items}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{e}")


def main():
    root = tk.Tk()
    app = ScoringApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
